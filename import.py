import collections
import mysql.connector
import settings
from openpyxl import load_workbook
from timer import Timer

'''
Created on Apr 22, 2017

@author: Phillip

SET FOREIGN_KEY_CHECKS=0;
TRUNCATE `card`;
TRUNCATE `coach`;
TRUNCATE `deck`;
TRUNCATE `inducement`;
TRUNCATE `match`;
TRUNCATE `match_inducement`;
TRUNCATE `match_player`;
TRUNCATE `match_player_skill`;
TRUNCATE `match_purchase`;
TRUNCATE `match_type`;
TRUNCATE `meta`;
TRUNCATE `player`;
TRUNCATE `player_skill`;
TRUNCATE `player_type`;
TRUNCATE `player_type_skill`;
TRUNCATE `player_type_skill_type_double`;
TRUNCATE `player_type_skill_type_normal`;
TRUNCATE `purchase`;
TRUNCATE `race`;
TRUNCATE `season`;
TRUNCATE `sessions`;
TRUNCATE `skill`;
TRUNCATE `skill_type`;
TRUNCATE `team`;
TRUNCATE `trophy`;

ALTER Table `card` AUTO_INCREMENT=1;
ALTER Table `coach` AUTO_INCREMENT=1;
ALTER Table `deck` AUTO_INCREMENT=1;
ALTER Table `inducement` AUTO_INCREMENT=1;
ALTER Table `match` AUTO_INCREMENT=1;
ALTER Table `match_inducement` AUTO_INCREMENT=1;
ALTER Table `match_player` AUTO_INCREMENT=1;
ALTER Table `match_player_skill` AUTO_INCREMENT=1;
ALTER Table `match_purchase` AUTO_INCREMENT=1;
ALTER Table `match_type` AUTO_INCREMENT=1;
ALTER Table `meta` AUTO_INCREMENT=1;
ALTER Table `player` AUTO_INCREMENT=1;
ALTER Table `player_type` AUTO_INCREMENT=1;
ALTER Table `player_type_skill` AUTO_INCREMENT=1;
ALTER Table `player_type_skill_type_double` AUTO_INCREMENT=1;
ALTER Table `player_type_skill_type_normal` AUTO_INCREMENT=1;
ALTER Table `purchase` AUTO_INCREMENT=1;
ALTER Table `race` AUTO_INCREMENT=1;
ALTER Table `season` AUTO_INCREMENT=1;
ALTER Table `skill` AUTO_INCREMENT=1;
ALTER Table `skill_type` AUTO_INCREMENT=1;
ALTER Table `team` AUTO_INCREMENT=1;
ALTER Table `trophy` AUTO_INCREMENT=1;

'''

TROPHIES = [
    ["Spike Bowl", "spike bowl.png"],
    ["Chaos Cup", "chaos cup.png"],
    ["Dungeon Bowl", "dungeon bowl.png"],
    ["Blood Bowl", "blood bowl.png"],
    ["Witch King Cup", "trophy.png"],
    ["Open Season", "trophy.png"]
]

SEASONS = [
    ["Season 1", "2017-01-01", "2017-03-31", "Spike Bowl", "Guardians of the Nile"],
    ["Season 2", "2017-04-01", "2017-06-30", "Chaos Cup", "Riverdale Ravagers"],
    ["Season 3", "2017-07-01", "2017-09-30", "Dungeon Bowl", "Rat Rat City"],
    ["Season 3.5", "2017-10-01", "2017-12-31", "Open Season", None],
    ["Season 4", "2018-01-01", "2018-03-31", "Witch King Cup", None],
    ["Season 5", "2018-05-21", "2018-08-31", "Blood Bowl", "Keepers Of The Nile"],
    ["Season 6", "2018-10-30", "2019-07-11", "Blood Bowl", "The Everlasting Hobstompers"],
    ["Season 6 Green League", "2018-10-30", "2019-07-11", "Open Season", None],
    ["Season 7", "2019-07-29", "2019-12-31", "Blood Bowl", None]
]

SKILL_TYPES = ["General", "Agility", "Passing", "Strength", "Mutation", "Extraordinary", "Improvement"]
SKILLS = [
    ["Accurate", "Passing", "The player may add 1 to the D6 roll when he passes."],
    ["Always Hungry", "Extraordinary", "The player is always ravenously hungry - and what's more, he'll eat absolutely anything! Should the player ever use the Throw Team-Mate skill, roll a D6 after he has finished moving, but before he throws his team-mate. On a 2+ continue with the throw. On a roll of 1 he attempts to eat the unfortunate team-mate! Roll the D6 again, a second 1 means that he successfully scoffs the team-mate down, which kills the team-mate without opportunity for recovery (Apothecaries, Regeneration or anything else cannot be used). If the team-mate had the ball it will scatter once from the team-mate's square. If the second roll is 2-6 the team-mate squirms free and the Pass Action is automatically treated as a fumbled pass. Fumble the player with the Right Stuff skill as normal."],
    ["Animosity", "Extraordinary", "A player with this skill does not like players from his team that are a different race than he is and will often refuse to play with them despite the coach's orders. If this player at the end of his Hand-off or Pass Action attempts to hand-off or pass the ball to a team-mate that is not the same race as the Animosity player, roll a D6. On a 2+, the pass/hand-off is carried out as normal. On a 1, the player refuses to give the ball to any team-mate except one of his own race. The coach may choose to change the target of the pass/hand-off to another team-mate of the same race as the Animosity player, however no more movement is allowed for the Animosity player, so the current Action may be lost for the turn"],
    ["Ball & Chain", "Extraordinary",
     "Players armed with a Ball & Chain can only take Move Actions. To move or Go For It, place the throw-in template over the player facing up or down the pitch or towards either sideline. Then roll a D6 and move the player one square in the indicated direction; no Dodge roll is required if you leave a tackle zone. If this movement takes the player off the pitch, he is beaten up by the crowd in the same manner as a player who has been pushed off the pitch. Repeat this process for each and every square of normal movement the player has. You may then GFI using the same process if you wish. If during his Move Action he would move into an occupied square then the player will throw a block following normal blocking rules against whoever is in that square, friend or foe (and it even ignores Foul Appearance!). A Prone or Stunned player in an occupied square is pushed back and an Armour roll is made to see if he is injured, instead of the block being thrown at him. The player must follow up if he will push back another player, and will then carry on with his move as described above. If the player is ever Knocked Down or Placed Prone, roll immediately for injury (no Armour roll is required). Stunned results for any Injury rolls for the Ball & Chain player are always treated as KO'd. A Ball & Chain player may use the Grab skill (as if a Block Action was being used) with his blocks (if he has learned it!). A Ball & Chain player may never use the Diving Tackle, Frenzy, Kick-Off Return, Leap, Pass Block or Shadowing skills."],
    ["Big Hand", "Mutation", "One of the player's hands has grown monstrously large, yet remained completely functional. The player ignores modifier(s) for enemy tackle zones or Pouring Rain weather when he attempts to pick up the ball."],
    ["Block", "General", "A player with the Block skill is proficient at knocking opponents down. The Block skill, if used, affects the results rolled with the Block dice, as explained in the Blocking rules."],
    ["Blood Lust", "Extraordinary",
     "Vampires must occasionally feed on the blood of the living. Immediately after declaring an Action with a Vampire, roll a d6: On a 2+ the Vampire can carry out the Action as normal. On a 1, however, the Vampire must feed on a Thrall team-mate or a spectator. The Vampire may continue with his declared Action or if he had declared a Block Action, he may take a Move Action instead. Either way, at the end of the declared Action, but before actually passing, handing off, or scoring, the vampire must feed. If he is standing adjacent to one or more Thrall team-mates (standing, prone or stunned), then choose one to bite and make an Injury roll on the Thrall treating any casualty roll as Badly Hurt. The injury will not cause a turnover unless the Thrall was holding the ball. Once the Vampire has bitten a Thrall he may complete his Action. Failure to bite a Thrall is a turnover and requires him to feed on a spectator - move the Vampire to the reserves box if he was still on the pitch. If he was holding the ball, it bounces from the square he occupied when he was removed and he will not score a touchdown if he was in the opposing end zone."],
    ["Bombardier", "Extraordinary",
     "A coach may choose to have a Bombardier who is not Prone or Stunned throw a bomb instead of taking any other Action with the player. This does not use the team's Pass Action for the turn. The bomb is thrown using the rules for throwing the ball (including weather effects and use of Hail Mary Pass), except that the player may not move or stand up before throwing it (he needs time to light the fuse!). Intercepted bomb passes are not turnovers. Fumbles or any bomb explosions that lead to a player on the active team being Knocked Down are turnovers. All skills that may be used when a ball is thrown may be used when a bomb is thrown also. A bomb may be intercepted or caught using the same rules for catching the ball, in which case the player catching it must throw it again immediately. This is a special bonus Action that takes place out of the normal sequence of play. A player holding the ball can catch or intercept and throw a bomb. The bomb explodes when it lands in an empty square or an opportunity to catch the bomb fails or is declined (i.e., bombs don't 'bounce'). If the bomb is fumbled it explodes in the bomb thrower's square. If a bomb lands in the crowd, it explodes with no effect. When the bomb finally does explode any player in the same square is Knocked Down, and players in adjacent squares are Knocked Down on a roll of 4+. Players can be hit by a bomb and treated as Knocked Down even if they are already Prone or Stunned. Make Armour and Injury rolls for any players Knocked Down by the bomb. Casualties caused by a bomb do not count for Star Player points."],
    ["Bone-head", "Extraordinary", "The player is not noted for his intelligence. Because of this you must roll a D6 immediately after declaring an Action for the player, but before taking the Action. On a roll of 1 he stands around trying to remember what it is he's meant to be doing. The player can't do anything for the turn, and the player's team loses the declared Action for the turn. (So if a Bone-head player declares a Blitz Action and rolls a 1, then the team cannot declare another Blitz Action that turn.) The player loses his tackle zones and may not catch, intercept or pass, assist another player on a block or foul, or voluntarily move until he manages to roll a 2 or better at the start of a future Action or the drive ends."],
    ["Break Tackle", "Strength", "The player may use his Strength instead of his Agility when making a Dodge roll. For example, a player with Strength 4 and Agility 2 would count as having an Agility of 4 when making a Dodge roll. This skill may only be used once per turn."],
    ["Catch", "Agility", "A player who has the Catch skill is allowed to re-roll the D6 if he fails a catch roll. It also allows the player to re-roll the D6 if he drops a hand-off or fails to make an interception."],
    ["Chainsaw", "Extraordinary",
     "A player armed with a chainsaw must attack with it instead of making a block as part of a Block or Blitz Action. When the chainsaw is used to make an attack, roll a D6 instead of the Block dice. On a roll of 2 or more the chainsaw hits the opposing player, but on a roll of 1 it 'kicks back' and hits the wielder instead! Make an Armour roll for the player hit by the chainsaw, adding 3 to the score. If the roll beats the victim's Armour value then the victim is Knocked Down and injured - roll on the Injury table. If the roll fails to beat the victim's Armour value then the attack has no effect. A player armed with a chainsaw may take a Foul Action, and adds 3 to the Armour roll, but must roll for kick back as described above. A running chainsaw is a dangerous thing to carry around, so if a player holding a chainsaw is Knocked Down for any reason, the opposing coach is allowed to add 3 to his Armour roll to see if the player was injured. However, blocking a player with a chainsaw is equally dangerous! If an opponent knocks himself over when blocking the chainsaw player then add 3 to his Armour roll. This skill may only be used once per turn (e.g., cannot be used with Frenzy or Multiple Block) and if used as part of a Blitz Action, the player cannot continue moving after using it. Casualties caused by a chainsaw player do not count for Star Player points."],
    ["Claws", "Mutation", "A player with this skill is blessed with a huge crab-like claw or razor sharp talons that make armour useless. When an opponent is Knocked Down by this player during a block, any Armour roll of 8 or more after modifications automatically breaks armour."],
    ["Dauntless", "General", "A player with this skill is capable of psyching himself up so he can take on even the very strongest opponent. The skill only works when the player attempts to block an opponent who is stronger than himself. When the skill is used, the coach of the player with the Dauntless skill rolls a D6 and adds it to his strength. If the total is equal to or lower than the opponent's Strength, the player must block using his normal Strength. If the total is greater, then the player with the Dauntless skill counts as having a Strength equal to his opponent's when he makes the block. The strength of both players is calculated before any defensive or offensive assists are added but after all other modifiers."],
    ["Decay", "Extraordinary", "Staying on the pitch is difficult when your rotting body is barely held together. When this player suffers a Casualty result on the Injury table, roll twice on the Casualty table (see page 25) and apply both results. The player will only ever miss one future match as a result of his injuries, even if he suffers two results with this effect. A successful Regeneration roll will heal both results."],
    ["Dirty Player", "General", "A player with this skill has trained long and hard to learn every dirty trick in the book. Add 1 to any Armour roll or Injury roll made by a player with this skill when they make a Foul as part of a Foul Action. Note that you may only modify one of the dice rolls, so if you decide to use Dirty Player to modify the Armour roll, you may not modify the Injury roll as well."],
    ["Disturbing Presence", "Mutation", "This player's presence is very disturbing, whether it is caused by a massive cloud of flies, sprays of soporific musk, an aura of random chaos or intense cold, or a pheromone that causes fear and panic. Regardless of the nature of this mutation, any player must subtract 1 from the D6 when they pass, intercept or catch for each opposing player with Disturbing Presence that is within three squares of them, even if the Disturbing Presence player is Prone or Stunned."],
    ["Diving Catch", "Agility", "The player is superb at diving to catch balls others cannot reach and jumping to more easily catch perfect passes. The player may add 1 to any catch roll from an accurate pass targeted to his square. In addition, the player can attempt to catch any pass, kick off or crowd throw-in, but not bouncing ball, that would land in an empty square in one of his tackle zones as if it had landed in his own square without leaving his current square. A failed catch will bounce from the Diving Catch player's square. If there are two or more players attempting to use this skill then they get in each other's way and neither can use it."],
    ["Diving Tackle", "Agility", "The player may use this skill after an opposing player attempts to dodge out of any of his tackle zones. The opposing player must subtract 2 from his Dodge roll for leaving the player's tackle zone. If a player is attempting to leave the tackle zone of several players that have the Diving Tackle skill, then only one of the opposing players may use Diving Tackle. Diving Tackle may be used on a re-rolled dodge if not declared for use on the first Dodge roll. Once the dodge is resolved but before any armour roll for the opponent (if needed), the Diving Tackle Player is Placed Prone in the square vacated by the dodging player but do not make an Armour or Injury roll for the Diving Tackle player."],
    ["Dodge", "Agility", "A player with the Dodge skill is adept at slipping away from opponents, and is allowed to re-roll the D6 if he fails to dodge out of any of an opposing player's tackle zones. However, the player may only re-roll one failed Dodge roll per turn. In addition, the Dodge skill, if used, affects the results rolled on the Block dice, as explained in the Blocking rules."],
    ["Dump-off", "Passing", "This skill allows the player to make a Quick Pass when an opponent declares that he will throw a block at him, allowing the player to get rid of the ball before he is hit. Work out the Dump-off pass before the opponent makes his block. The normal throwing rules apply, except that neither team's turn ends as a result of the throw, whatever it may be. After the throw is worked out your opponent completes the block, and then carries on with his turn. Dump-off may not be used on the second block from an opponent with the Frenzy skill or in conjunction with the Bombardier or Throw Team-Mate skills."],
    ["Extra Arms", "Mutation", "A player with one or more extra arms may add 1 to any attempt to pick up, catch or intercept."],
    ["Fan Favorite", "Extraordinary", "The fans love seeing this player on the pitch so much that even the opposing fans cheer for your team. For each player with Fan Favorite on the pitch your team receives an additional +1 FAME modifier (see page 18) for any Kick-Off table results, but not for the Winnings roll."],
    ["Fend", "General", "This player is very skilled at holding off would-be attackers. Opposing players may not follow-up blocks made against this player even if the Fend player is Knocked Down. The opposing player may still continue moving after blocking if he had declared a Blitz Action."],
    ["Foul Appearance", "Mutation", "The player's appearance is so horrible that any opposing player that wants to block the player (or use a special attack that takes the place of a block) must first roll a D6 and score 2 or more. If the opposing player rolls a 1 he is too revolted to make the block and it is wasted (though the opposing team does not suffer a turnover)."],
    ["Frenzy", "General", "A player with this skill is a slavering psychopath who attacks his opponents in an uncontrollable rage. Unless otherwise overridden, this skill must always be used. When making a block, a player with this skill must always follow up if he can. If a 'Pushed' or 'Defender Stumbles' result was chosen, the player must immediately throw a second block against the same opponent so long as they are both still standing and adjacent. If possible, the player must also follow up this second block. If the frenzied player is performing a Blitz Action then he must pay a square of Movement and must make the second block unless he has no further normal movement and cannot Go For It again."],
    ["Grab", "Strength", "A player with this skill uses his great strength and prowess to grab his opponent and throw him around. To represent this, only while making a Block Action, if his block results in a push back he may choose any empty square adjacent to his opponent to push back his opponent. When making a Block or Blitz Action, Grab and Side Step will cancel each other out and the standard pushback rules apply. Grab will not work if there are no empty adjacent squares. A player with the Grab skill can never learn or gain the Frenzy skill through any means. Likewise, a player with the Frenzy skill can never learn or gain the Grab skill through any means."],
    ["Guard", "Strength", "A player with this skill assists an offensive or defensive block even if he is in another player's tackle zone. This skill may not be used to assist a foul."],
    ["Hail Mary Pass", "Passing", "The player may throw the ball to any square on the playing pitch, no matter what the range: the range ruler is not used. Roll a D6. On a roll of 1 the player fumbles the throw, and the ball will bounce once from the thrower's square. On a roll of 2-6, the player may make the pass. The Hail Mary pass may not be intercepted, but it is never accurate - the ball automatically misses and scatters three squares. Note that if you are lucky, the ball will scatter back into the target square! This skill may not be used in a Blizzard or with the Throw Team-Mate skill."],
    ["Horns", "Mutation", "A player with Horns may use them to butt an opponent. Horns adds 1 to the player's Strength for any block(s) he makes during a Blitz Action."],
    ["Hypnotic Gaze", "Extraordinary", "The player has a powerful telepathic ability that he can use to stun an opponent into immobility. The player may use hypnotic gaze at the end of his Move Action on one opposing player who is in an adjacent square. Make an Agility roll for the player with hypnotic gaze, with a -1 modifier for each opposing tackle zone on the player with hypnotic gaze other than the victim's. If the Agility roll is successful, then the opposing player loses his tackle zones and may not catch, intercept or pass the ball, assist another player on a block or foul, or move voluntarily until the start of his next Action or the drive ends. If the roll fails, then the hypnotic gaze has no effect."],
    ["Juggernaut", "Strength", "A player with this skill is virtually impossible to stop once he is in motion. If this player takes a Blitz Action, the opposing player may not use his Fend, Stand Firm or Wrestle skills against the Juggernaut player's blocks. The Juggernaut player may also choose to treat a 'Both Down' result as if a 'Pushed' result has been rolled instead for blocks he makes during a Blitz Action."],
    ["Jump Up", "Agility", "A player with this skill is able to quickly get back into the game. If the player declares any Action other than a Block Action he may stand up for free without paying the three squares of movement. The player may also declare a Block Action while Prone which requires an Agility roll with a +2 modifier to see if he can complete the Action. A successful roll means the player can stand up for free and block an adjacent opponent. A failed roll means the Block Action is wasted and the player may not stand up."],
    ["Kick", "General", "The player is an expert at kicking the ball and can place the kick with great precision. In order to use this skill the player must be set up on the pitch when his team kicks off. The player may not be set up in either wide zone or on the line of scrimmage. Only if all these conditions are met is the player then allowed to take the kick-off. Because his kick is so accurate, you may choose to halve the number of squares that the ball scatters on kick-off, rounding any fractions down (i.e., 1 = 0, 2-3 = 1, 4- 5 = 2, 6 = 3)."],
    ["Kick-Off Return", "General", "A player on the receiving team that is not on the Line of Scrimmage or in an opposing tackle zone may use this skill when the ball has been kicked. It allows the player to move up to 3 squares after the ball has been scattered but before rolling on the Kick-Off table. Only one player may use this skill each kick-off. This skill may not be used for a touchback kick-off and does not allow the player to cross into the opponent's half of the pitch."],
    ["Leader", "Passing", "The player is a natural leader and commands the rest of the team from the back-field as he prepares to throw the ball. A team with one or more players with the Leader skill may take a single Leader Re-roll counter and add it to their team re-rolls at the start of the game and at half time after any Master Chef rolls. The Leader re-roll is used exactly the same in every way as a normal Team re-roll with all the same restrictions. In addition, the Leader re-roll may only be used so long as at least one player with the Leader skill is on the pitch - even if they are Prone or Stunned! Rerolls from Leader may be carried over into Overtime if not used, but the team does not receive a new Leader re-roll at the start of Overtime."],
    ["Leap", "Agility", "A player with the Leap skill is allowed to jump to any empty square within 2 squares even if it requires jumping over a player from either team. Making a leap costs the player two squares of movement. In order to make the leap, move the player to any empty square 1 to 2 squares from his current square and then make an Agility roll for the player. No modifiers apply to this D6 roll unless he has Very Long Legs. The player does not have to dodge to leave the square he starts in. If the player successfully makes the D6 roll then he makes a perfect jump and may carry on moving. If the player fails the Agility roll then he is Knocked Down in the square that he was leaping to, and the opposing coach makes an Armour roll to see if he was injured. A player may only use the Leap skill once per turn."],
    ["Loner", "Extraordinary", "Loners, through inexperience, arrogance, animal ferocity or just plain stupidity, do not work well with the rest of the team. As a result, a Loner may use team re-rolls but has to roll a D6 first. On a roll of 4+, he may use the team re-roll as normal. On a roll of 1-3 the original result stands without being re-rolled but the team re-roll is lost (i.e., used)."],
    ["Mighty Blow", "Strength", "Add 1 to any Armour or Injury roll made by a player with this skill when an opponent is Knocked Down by this player during a block. Note that you only modify one of the dice rolls, so if you decide to use Mighty Blow to modify the Armour roll, you may not modify the Injury roll as well. Mighty Blow cannot be used with the Stab or Chainsaw skills."],
    ["Multiple Block", "Strength", "At the start of a Block Action a player who is adjacent to at least two opponents may choose to throw blocks against two of them. Make each block in turn as normal except that each defender's strength is increased by 2. The player cannot follow up either block when using this skill, so Multiple Block can be used instead of Frenzy, but both skills cannot be used together. To have the option to throw the second block the player must still be on his feet after the first block."],
    ["Monstrous Mouth", "Extraordinary", "A player with a Monstrous Mouth is allowed to re-roll the D6 if they fail a Catch roll. It also allows the player to re-roll the D6 if they drop a hand-off or fail to make an interception. In addition, the Strip Ball skill will not work against a player with a Monstrous Mouth."],   
    ["Nerves of Steel", "Passing", "The player ignores modifiers for enemy tackle zones when he attempts to pass, catch or intercept."],
    ["No Hands", "Extraordinary", "The player is unable to pick up, intercept or carry the ball and will fail any catch roll automatically, either because he literally has no hands or because his hands are full. If he attempts to pick up the ball then it will bounce, and will cause a turnover if it is his team's turn."],
    ["Nurgle's Rot", "Extraordinary", "This player has a horrible infectious disease which spreads when he kills an opponent during a Block, Blitz or Foul Action. Instead of truly dying, the infected opponent becomes a new rookie Rotter. To do so, the opponent must have been removed from the roster during step 2.1 of the Post-game sequence, his Strength cannot exceed 4, and he cannot have the Decay, Regeneration or Stunty skills. The new Rotter can be added to the Nurgle team for free during step 5 of Updating Your Team Roster (see page 29) if the team has an open Roster slot. This new Rotter still counts at full value towards the total value of the Nurgle team."],
    ["Pass", "Passing", "A player with the Pass skill is allowed to re-roll the D6 if he throws an inaccurate pass or fumbles."],
    ["Pass Block", "General",
     "A player with this skill is allowed to move up to three squares when the opposing coach announces that one of his players is going to pass the ball (but not a bomb). The opposing coach may not change his mind about passing once Pass Block's use is declared. The move is made out of sequence, after the range has been measured, but before any interception attempts have been made. A player may not make the move unless able to reach a legal destination and may not follow a route that would not allow them to reach a legal destination. A legal destination puts the player in a position to attempt an interception, an empty square that is the target of the pass, or with his tackle zone on the thrower or catcher. The player may not stop moving until he has reached a legal destination, has been held fast by Tentacles or has been Knocked Down. The special move is free, and in no way affects the player's ability to move in a subsequent action. The move is made using all of the normal rules and skills (for example, having to dodge in order to leave opposing players' tackle zones.) Players with Pass Block may use this skill against a Dump Off pass. If a player performing a Pass Block in their own turn is Knocked Down then this is a turnover, no other players may perform Pass Block moves, and your turn ends as soon as the results of the pass and the block are resolved."],
    ["Piling On", "Strength", "The player may use this skill after he has made a block as part of one of his Block or Blitz Actions, but only if the Piling On player is currently standing adjacent to the victim and the victim was Knocked Down. You may re-roll the Armour roll or Injury roll for the victim. The Piling On player is Placed Prone in his own square -- it is assumed that he rolls back there after flattening his opponent (do not make an Armour roll for him as he has been cushioned by the other player!). Piling On does not cause a turnover unless the Piling On player is carrying the ball. Piling On cannot be used with the Stab or Chainsaw skills."],
    ["Prehensile Tail", "Mutation", "The player has a long, thick tail which he can use to trip up opposing players. To represent this, opposing players must subtract 1 from the D6 roll if they attempt to dodge out of any of the player's tackle zones."],
    ["Pro", "General", "A player with this skill is a hardened veteran. Such players are called professionals or Pros by other Blood Bowl players because they rarely make a mistake. Once per turn, a Pro is allowed to re-roll any one dice roll he has made other than Armour, Injury or Casualty, even if he is Prone or Stunned. However, before the re-roll may be made, his coach must roll a D6. On a roll of 4, 5 or 6 the re-roll may be made. On a roll of 1, 2 or 3 the original result stands and may not be re-rolled with a skill or team re-roll; however you can re-roll the Pro roll with a Team re-roll."],
    ["Really Stupid", "Extraordinary",
     "This player is without doubt one of the dimmest creatures to ever take to a Blood Bowl pitch (which, considering the IQ of most other players, is really saying something!). Because of this you must roll a D6 immediately after declaring an Action for the player, but before taking the Action. If there are one or more players from the same team standing adjacent to the Really Stupid player's square, and who aren't Really Stupid, then add 2 to the D6 roll. On a result of 1-3 he stands around trying to remember what it is he's meant to be doing. The player can't do anything for the turn, and the player's team loses the declared Action for that turn (for example, if a Really Stupid player declares a Blitz Action and fails the Really Stupid roll, then the team cannot declare another Blitz Action that turn). The player loses his tackle zones and may not catch, intercept or pass the ball, assist another player on a block or foul, or voluntarily move until he manages to roll a successful result for a Really Stupid roll at the start of a future Action or the drive ends."],
    ["Regeneration", "Extraordinary", "If the player suffers a Casualty result on the Injury table, then roll a D6 for Regeneration after the roll on the Casualty table and after any Apothecary roll, if allowed. On a result of 1-3, the player suffers the result of this injury. On a 4-6, the player will heal the injury after a short period of time to 're-organise' himself, and is placed in the Reserves box instead. Regeneration rolls may not be re-rolled. Note that an opposing player still earns Star Player points as normal for inflicting a Casualty result on a player with this skill, even if the result doesn't affect the player in the normal way."],
    ["Right Stuff", "Extraordinary", "A player with the Right Stuff skill can be thrown by another player from his team who has the Throw Team-Mate skill. See the Throw Team-Mate skill entry below for details of how the player is thrown. When a player with this skill is thrown or fumbled and ends up in an unoccupied square, he must make a landing roll unless he landed on another player during the throw. A landing roll is an Agility roll with a -1 modifier for each opposing player's tackle zone on the square he lands in. If he passes the roll he lands on his feet. If the landing roll is failed or he landed on another player during the throw he is Placed Prone and must pass an Armour roll to avoid injury. If the player is not injured during his landing he may take an Action later this turn if he has not already done so. A failed landing roll or landing in the crowd does not cause a turnover, unless he was holding the ball."],
    ["Safe Throw", "Passing", "This player is an expert at throwing the ball in a way that makes it even more difficult for any opponent to intercept it. If a pass made by this player is ever intercepted then the Safe Throw player may make an unmodified Agility roll. If successful, the interception is cancelled out and the passing sequence continues as normal. In addition, if this player fumbles a pass of a ball (not a bomb or team-mate) on any roll other than a natural 1, he manages to keep hold of the ball instead of suffering a fumble and the team does not suffer a turnover."],
    ["Secret Weapon", "Extraordinary", "Some players are armed with special pieces of equipment that are called 'secret weapons.' Although the Blood Bowl rules specifically ban the use of any weapons, the game has a long history of teams trying to get weapons of some sort onto the pitch. Nonetheless, the use of secret weapons is simply not legal, and referees have a nasty habit of sending off players that use them. Once a drive ends that this player has played in at any point, the referee orders the player to be sent off to the dungeon to join players that have been caught committing fouls during the match, regardless of whether the player is still on the pitch or not."],
    ["Shadowing", "General", "The player may use this skill when a player performing an Action on the opposing team moves out of any of his tackle zones for any reason. The opposing coach rolls 2D6 adding his own player's movement allowance and subtracting the Shadowing player's movement allowance from the score. If the final result is 7 or less, the player with Shadowing may move into the square vacated by the opposing player. He does not have to make any Dodge rolls when he makes this move, and it has no effect on his own movement in his own turn. If the final result is 8 or more, the opposing player successfully avoids the Shadowing player and the Shadowing player may not move into the vacated square. A player may make any number of shadowing moves per turn. If a player has left the tackle zone of several players that have the Shadowing skill, then only one of the opposing players may attempt to shadow him."],
    ["Side Step", "Agility", "A player with this skill is an expert at stepping neatly out of the way of an attacker. To represent this ability, his coach may choose which square the player is moved to when he is pushed back, rather than the opposing coach. Furthermore, the coach may choose to move the player to any adjacent square, not just the three squares shown on the Push Back diagram. Note that the player may not use this skill if there are no open squares on the pitch adjacent to this player. Note that the coach may choose which square the player is moved to even if the player is Knocked Down after the push back."],
    ["Sneaky Git", "Agility", "This player has the quickness and finesse to stick the boot to a downed opponent without drawing a referee's attention unless he hears the armour crack. During a Foul Action a player with this skill is not ejected for rolling doubles on the Armour roll unless the Armour roll was successful."],
    ["Sprint", "Agility", "The player may attempt to move up to three extra squares rather than the normal two when Going For It (see page 20). His coach must still roll to see if the player is Knocked Down in each extra square he enters."],
    ["Stab", "Extraordinary", "A player with this skill is armed with something very good at stabbing, slashing or hacking up an opponent, like sharp fangs or a trusty dagger. This player may attack an opponent with his stabbing attack instead of throwing a block. Make an unmodified Armour roll (except for Stakes) for the victim. If the score is less than or equal to the victim's Armour value then the attack has no effect. If the score beats the victim's Armour value then he has been wounded and an Injury roll must be made. This Injury roll ignores all modifiers from any source - including Niggling injuries. If Stab is used as part of a Blitz Action, the player cannot continue moving after using it. Casualties caused by a stabbing attack do not count for Star Player points."],
    ["Stakes", "Extraordinary", "This player is armed with special stakes that are blessed to cause extra damage to the Undead and those that work with them. This player may add 1 to the Armour roll when they make a Stab attack against any player playing for a Khemri, Necromantic, Undead or Vampire team."],
    ["Stand Firm", "Strength", "A player with this skill may choose to not be pushed back as the result of a block. He may choose to ignore being pushed by 'Pushed' results, and to have 'Knock-down' results knock the player down in the square where he started. If a player is pushed back into a player using Stand Firm then neither player moves."],
    ["Strip Ball", "General", "When a player with this skill blocks an opponent with the ball, applying a 'Pushed' or 'Defender Stumbles' result will cause the opposing player to drop the ball in the square that they are pushed to, even if the opposing player is not Knocked Down."],
    ["Strong Arm", "Strength", "The player may add 1 to the D6 when he passes to Short, Long or Long Bomb range."],
    ["Stunty", "Extraordinary", "The player is so small that they are very difficult to tackle because they can duck underneath opposing players' outstretched arms and run between their legs. On the other hand, Stunty players are just a bit too small to throw the ball very well, and are easily injured. To represent these things a player with the Stunty skill may ignore any enemy tackle zones on the square he is moving to when he makes a Dodge roll (i.e., they always end up with a +1 Dodge roll modifier), but must subtract 1 from the roll when he passes. In addition, this player treats a roll of 7 and 9 on the Injury table after any modifiers have been applied as a KO'd and Badly Hurt result respectively, rather than the normal results. Stunties that are armed with a Secret Weapon are not allowed to ignore enemy tackle zones, but still suffer the other penalties."],
    ["Sure Feet", "Agility", "The player may re-roll the D6 if he is Knocked Down when trying to Go For It. A player may only use the Sure Feet skill once per turn."],
    ["Sure Hands", "General", "A player with the Sure Hands skill is allowed to re-roll the D6 if he fails to pick up the ball. In addition, the Strip Ball skill will not work against a player with this skill."],
    ["Swoop", "Extraordinary", "This player is equipped with a rudimentary set of wings, either natural or engineered, allowing them to glide through the air (rather than plummeting gracelessly) if they are thrown by a team-mate. If a player with Swoop is thrown by a player with the Throw Team-mate skill, the Throw-in template is used instead of the Scatter template to see where they land. Each time the player scatters, their coach places the Throw-in template over the player facing up or down the pitch or towards either sideline. Then they roll a D6 and move the player one square in the indicated direction. In addition, when rolling to see whether the player lands on their feet (as per the Right Stuff skill), add 1 to the result. When a player with both the Swoop and Stunty skills dodges, they do not ignore any modifiers for enemy tackle zones on the square they are moving to - the presence of a large pair of wings negates any benefit they would gain from being small and slippery."],
    ["Tackle", "General", "Opposing players who are standing in any of this player's tackle zones are not allowed to use their Dodge skill if they attempt to dodge out of any of the player's tackle zones, nor may they use their Dodge skill if the player throws a block at them and uses the Tackle skill."],
    ["Take Root", "Extraordinary", "Immediately after declaring an Action with this player, roll a D6. On a 2 or more, the player may take his Action as normal. On a 1, the player \"takes root\", and his MA is considered 0 until a drive ends, or he is Knocked Down or Placed Prone (and no, players from his own team may not try and block him in order to try to knock him down!). A player that has taken root may not Go For It, be pushed back for any reason, or use any skill that would allow him to move out of his current square or be Placed Prone. The player may block adjacent players without following-up as part of a Block Action however if a player fails his Take Root roll as part of a Blitz Action he may not block that turn (he can still roll to stand up if he is Prone)."],
    ["Tentacles", "Mutation", "The player may use this skill when an opposing player attempts to dodge or leap out of any of his tackle zones. The opposing coach rolls 2D6 adding his player's ST and subtracting the Tentacles player's ST from the score. If the final result is 5 or less, then the moving player is held firm, and his Action ends immediately. If a player attempts to leave the tackle zone of several players that have the Tentacles skill, only one may attempt to grab him with Tentacles."],
    ["Thick Skull", "Strength", "This player treats a roll of 8 on the Injury table, after any modifiers have been applied, as a Stunned result rather than a KO'd result. This skill may be used even if the player is Prone or Stunned."],
    ["Throw Team-Mate", "Extraordinary",
     "A player with this skill has the ability to throw a player from the same team instead of the ball! (This includes the ball if the player thrown already has it!) The throwing player must end the movement of his Pass Action standing next to the intended team-mate to be thrown, who must have the Right Stuff skill and be standing. The pass is worked out exactly the same as if the player with Throw Team-Mate was passing a ball, except the player must subtract 1 from the D6 roll when he passes the player, fumbles are not automatically turnovers, and Long Pass or Long Bomb range passes are not possible. In addition, accurate passes are treated instead as inaccurate passes thus scattering the thrown player three times as players are heavier and harder to pass than a ball. The thrown player cannot be intercepted. A fumbled team-mate will land in the square he originally occupied. If the thrown player scatters off the pitch, he is beaten up by the crowd in the same manner as a player who has been pushed off the pitch. If the final square he scatters into is occupied by another player, treat the player landed on as Knocked Down and roll for Armour (even if already Prone or Stunned), and then the player being thrown will scatter one more square. If the thrown player would land on another player, continue to scatter the thrown player until he ends up in an empty square or off the pitch (i.e., he cannot land on more than one player). See the Right Stuff entry to see if the player lands on his feet or head-down in a crumpled heap!"],
    ["Timmm-ber!", "Extraordinary", "This player spends so much time on the floor that their teammates have developed a knack for helping them up. If a player with this skill attempts to stand up after being knocked over, other players from their team can assist if they are adjacent, standing, and not in any enemy tackle zones. Each player that assists in this way adds 1 to the result of the dice roll to see whether the player stands up, but remember that a 1 is always a failure, no matter how many players are helping! Assisting a player to stand up does not count as an Action, and a player can assist regardless of whether they have taken an Action."],
    ["Titchy", "Extraordinary", "Titchy players tend to be even smaller and more nimble than other Stunty players. To represent this, the player may add 1 to any Dodge roll he attempts. On the other hand, while opponents do have to dodge to leave any of a Titchy player's tackle zones, a Titchy player is so small that he does not exert a -1 modifier when opponents dodge into any of his tackle zones."],
    ["Two Heads", "Mutation", "Having two heads enables this player to watch where he is going and the opponent trying to make sure he does not get there at the same time. Add 1 to all Dodge rolls the player makes."],
    ["Very Long Legs", "Mutation", "The player is allowed to add 1 to the D6 roll whenever he attempts to intercept or uses the Leap skill. In addition, the Safe Throw skill may not be used to affect any Interception rolls made by this player."],
    ["Wild Animal", "Extraordinary", "Wild Animals are uncontrollable creatures that rarely do exactly what a coach wants of them. In fact, just about all you can really rely on them to do is lash out at opposing players that move too close to them! To represent this, immediately after declaring an Action with a Wild Animal, roll a D6, adding 2 to the roll if taking a Block or Blitz Action. On a roll of 1-3, the Wild Animal does not move and roars in rage instead, and the Action is wasted."],
    ["Weeping Dagger", "Extraordinary", "This player keeps a warpstone-tainted dagger hidden in their kit, and is an expert at keeping it out of the referee's sight! If this player inflicts a casualty during a block, and the result of the Casualty roll is 11-38 (Badly Hurt) after any re-rolls, roll a D6. On a result of 4 or more, the opposing player must miss their next game. If you are not playing a league, a Weeping Dagger has no effect on the game."],
    ["Wrestle", "General", "The player is specially trained in grappling techniques. This player may use Wrestle when he blocks or is blocked and a 'Both Down' result on the Block dice is chosen by either coach. Instead of applying the 'Both Down' result, both players are wrestled to the ground. Both players are Placed Prone in their respective squares even if one or both have the Block skill. Do not make Armour rolls for either player. Use of this skill does not cause a turnover unless the active player was holding the ball."],
    ["+MA", "Improvement", ""],
    ["+AV", "Improvement", ""],
    ["+AG", "Improvement", ""],
    ["+ST", "Improvement", ""]
]
RACES = [
    ["Amazon", "Long ago, driven by a desire for adventure, the Valkyries of the Norse settlement in Lustria sailed away from their men-folk and founded a colony deep within the estuary of the river Amaxon. Now these ferocious warriors have taken to the Blood Bowl pitch - and Nuffle save those who dare play against them!"],
    ["Chaos Chosen", "Chaos Chosen teams are not noted for the subtlety or originality of their game play. A simple drive up the centre of the pitch, maiming and injuring as many opposing players as possible, is about the limit of their game plan. They rarely, if ever, worry about such minor considerations like picking up the ball and scoring touchdowns - not while there are any players left alive on the opposing team, anyway."],
    ["Chaos Dwarf", "Chaos Dwarfs are the twisted descendants of Dwarf explorers who have been terribly affected by the forces of Chaos, turning them into evil, self-centred creatures. In one way, however, they haven't changed at all - they still love playing Blood Bowl! Chaos Dwarfs are not very numerous and make great use of sneaky Hobgoblin slaves to perform all kinds of tasks, including playing on their Blood Bowl teams."],
    ["Dark Elf", "Evil beyond belief, skilled without doubt, the Dark Elves take to the pitch to show the world their superiority. Dark Elf teams prefer a malevolent and spiteful running game over the passing of their goodly cousins. Backed up by the ruthless Witch Elves and dangerous assassins, a Dark Elf team has all the tools to power through rather than around any opposition line."],
    ["Dwarf", "Dwarfs seem to be ideal Blood Bowl players, being compact, tough, well-armoured and having a stubborn knack of refusing to die! Most successful Dwarf teams work to the principle that if they can take out all the other team's potential scorers, and wear down the rest, then there won't be anybody left to stop them scoring the winning touchdowns!"],
    ["Elven Union", "When the NAF collapsed, many Elven teams were left penniless. Those teams that have survived the fallout are not as rich as the High Elf teams nor as well equipped, but they sure know how to play the game. Sporting facemasks and mohawks, they take to the pitch to relive the glory days they once played in."],
    ["Goblin", "A Goblin team's game plan owes much more to hope than potential. Goblins can make quite good catchers because they are small and agile, but the art of throwing is sadly lost to them, while the chances of their blocking anything larger than a Halfling are remote to say the least. Still, this never seems to bother Goblin players, and occasionally the use of a particularly devious secret weapon will even allow a Goblin team to win a match."],
    ["Halfling", "The technical deficiency of Halfling teams is legendary. They're too short to throw or catch, they run at half pace, and the whole team can spend all afternoon trying to block an Ogre without any chance of success. Most Halfling coaches try to make up for quality with quantity. After all, if you can get half a dozen players in the opposing team's End Zone and, by some miracle, manage to end up with the ball, then there is a small chance that one or two of them won't be jelly by the time you throw the thing."],
    ["High Elf", "The Elven Kingdom sponsored High Elf teams, feature a dangerous passing game and some of the most arrogant players you will find. Rich beyond the dreams of most teams, the High Elves often feature many Princes and noble born Elves on the team and what they cannot beat, they'll buy."],
    ["Human", "Although Human teams do not have the individual strengths or outstanding abilities available to other races, they do not suffer from any outstanding weakness either. This makes Human teams extremely flexible, equally at home running the ball, passing it, or ignoring it and pounding the opposition into the turf instead!"],
    ["Khemri Tomb Kings", "Over 8,000 years ago, the Khemri played the first games of Blood Bowl against the Slann. But, as the Kingdom died off, so did the game until its rediscovery. And as the game returned, it was inevitable that the ancient players and stars of the Khemri would return to the pitch they once played on."],
    ["Lizardman", "The Mage-Priests foretold the game of Blood Bowl thousands of years before it was discovered by the Dwarf Roze-El. So it is no surprise that the Lizardmen play Blood Bowl. Providing an odd blend of dexterity and strength, the Lustrian team can almost last the distance against a power team such as Chaos, while remaining able to pull off the running plays of the Skaven."],
    ["Necromantic Horror", "The damned and the cursed do not always lurk in the forests or in the graveyards of the Old World. Sometimes they come together, forming a group to hunt those more fortunate of souls. Finding relief in crazed outbursts of terrible violence, these groups do the best they can to ease their suffering - they pop off for a nice game of Blood Bowl."],
    ["Norse", "Norse teams have a well deserved reputation for ferocity both on and off the playing pitch. The Norse that takes up Blood Bowl is a truly unedifying specimen, interested only in beer, women and song off the playing pitch, and beer, women and bloody carnage while on it!"],
    ["Nurgle", "Nurgle teams are a form of Chaos team whose players worship the god Nurgle. Nurgle is the Chaos god of corruption and disease, and he rewards his players by granting them a rather unpleasant disease known as Nurgle's Rot. The fact that Nurgle teams smell awful is assumed rather than proven. True, they are all made up of semi decomposed flesh surrounded by swarms of flies, but by the time anyone gets close enough to get a really accurate whiff, he has inevitably caught one of Nurgle's nasty diseases, and he usually dies before he can suggest a new personal hygiene regime."],
    ["Ogre", "Ogre teams have existed since the forming of the NAF and have even had some success such as winning the XV Blood Bowl. However, as any right-minded person will tell you, having more than one Ogre in the same place at the same time is a disaster in the making. The key to an Ogre team is the Snotlings. If they are close enough to jab an Ogre in the leg to remind him that they are playing in a match then you may have the makings of a team."],
    ["Orc", "Orcs have been playing Blood Bowl since the game was invented, and Orc teams such as the Gouged Eye and Severed Heads are amongst the best in the league. Orc teams are tough and hard-hitting, grinding down the opposition's line to create gaps for their excellent Orc Blitzers to exploit."],
    ["Skaven", "They may not be all that strong, they certainly aren't tough, but boy oh boy are Skaven fast! Many an opponent has been left in the starting blocks as fast-moving Skaven players scamper through a gap in the line and run in for a lightning fast touchdown."],
    ["Shambling Undead", "In the Old World the dead do not rest easy. Vampires lurk in haunted castles, Necromancers seek to escape death by searching for forbidden knowledge, the Liche-lords rule over legions of corpses, and on the Blood Bowl field players who died long ago return to the scenes of their former glory and play Blood Bowl once again..."],
    ["Vampire", "Although Vampire teams include a number of extremely capable players, they are let down by the unreliability of the Vampires. While they should be concentrating on the game, their attention often wanders to their hunger and before you know it they are off for a quick bite!"],
    ["Wood Elf", "For Wood Elves the Long pass is everything, even more so than their High Elf cousins, and all of their effort goes into being an expert at throwing or receiving. No Wood Elf worth his salt is going to be weighed down by extra Armour and be forced to lurk about, attempting to knock opposing players over. Instead they rely on their natural athletic ability to keep them out of trouble, which is normally enough - it takes a very agile or lucky opponent to lay a hand on a Wood Elf!"],
    ["Chaos Pact", "Chaos Pact teams are a mix of evil and chaotic races. The Marauders while enthusiastic have to be coached to fill the different needs of the team while other races provide the muscle and fineness to support the Marauders. However due to the arrogance, stupidity, or animalistic nature of the team members, it is rare to see a well organized and effective Chaos Pact team. The Chaos All-Stars are the best example of how great this team can be with the right coach."],
    ["Slann", "The Slann team is an ancient race of space travellers stranded on our planets many ages ago. After realizing that rescue was never coming they settled down and began ordering the Lizardmen around as their leaders. While most Slann prefer to become fat and lazy lording over the Lizardmen, a few of younger and more energetic members enjoy travelling the realm and playing Blood Bowl. While the Slann have no passing game to speak of, their ability to leap, dive, and intercept are second to none."],
    ["Underworld Denizens", "On occasion the Skaven and Goblins living below all the hated races walking above in the sun team together to form Blood Bowl teams. The Underworld Creepers are the best known and most successful of these Underworld teams to date. However these teams often have very poor records as they spend most of their time infighting and blaming each other for the errors for each play. The one feature of this team that makes many fans attend is that the players sleep and bathe in Warpstone (and some eat it). While this kills off most of the potential players before they ever join a team, the ones that do survive often develop fascinating mutations."],
    ["Bretonnian", "In the fair land of Bretonnia arrogant Bretonnian nobles and their most trusted yeomen are questing for an alternative grail - the coveted Bloodweiser trophy. All too convinced of their own skill, the young knights fill out their team with lineman levy, drafted from the many local and remarkably incompetent all-peasant teams."],
    ["Daemons of Khorne", "Blood for the Blood God! This is the chant of the frenzied fans of the Khorne team demanding blood for Khorne and the players are happy to give it to them often by launching opponents off the pitch for sacrificial mayhem. Uncontrollable rage turns the team into virtually unstoppable blitzers making the sidelines very dangerous. The Lord of Rage blesses the team with his Daemons along with the most feared raving blitzer monstrosity to walk into Blood Bowl: the Bloodthirster. Khorne desires, nay expects, the field and a victory to be his, and bloody..."],
    ["Simyin", "The Simyin handle the ball well with all the Extra Arms. The Silverback and Gorillas break up opposing formations with mass-Grab, while the Chimpanzees make excellent ball-hunters. Bonobos are versatile and the Orangutan can develop into a powerful thrower."]
]

PLAYER_TYPES = [
    ["Amazon", 16, "Tribal Linewoman", 50, 6, 3, 3, 7, ["Dodge"], "G", "ASP"],
    ["Amazon", 2, "Eagle Warrior Thrower", 70, 6, 3, 3, 7, ["Dodge", "Pass"], "GP", "AS"],
    ["Amazon", 2, "Piranha Warrior Catcher", 70, 6, 3, 3, 7, ["Dodge", "Catch"], "GA", "SP"],
    ["Amazon", 4, "Koka Kalim Blitzer", 90, 6, 3, 3, 7, ["Dodge", "Block"], "GS", "AP"],

    ["Chaos Chosen", 16, "Beastman", 60, 6, 3, 3, 8, ["Horns"], "GSM", "AP"],
    ["Chaos Chosen", 4, "Chaos Warrior", 100, 5, 4, 3, 9, [], "GSM", "AP"],
    ["Chaos Chosen", 1, "Minotaur", 150, 5, 5, 2, 8, ["Loner", "Frenzy", "Horns", "Mighty Blow", "Thick Skull", "Wild Animal"], "SM", "GAP"],

    ["Chaos Dwarf", 16, "Hobgoblin", 40, 6, 3, 3, 7, [], "G", "ASP"],
    ["Chaos Dwarf", 6, "Chaos Dwarf", 70, 4, 3, 2, 9, ["Block", "Tackle", "Thick Skull"], "GS", "APM"],
    ["Chaos Dwarf", 2, "Bull Centaur", 130, 6, 4, 2, 9, ["Sprint", "Sure Feet", "Thick Skull"], "GS", "AP"],
    ["Chaos Dwarf", 1, "Enslaved Minotaur", 150, 5, 5, 2, 8, ["Loner", "Frenzy", "Horns", "Mighty Blow", "Thick Skull", "Wild Animal"], "S", "GAPM"],

    ["Dark Elf", 16, "Lineman", 70, 6, 3, 4, 8, [], "GA", "SP"],
    ["Dark Elf", 2, "Runner", 80, 7, 3, 4, 7, ["Dump-off"], "GAP", "S"],
    ["Dark Elf", 2, "Assassin", 90, 6, 3, 4, 7, ["Shadowing", "Stab"], "GA", "SP"],
    ["Dark Elf", 4, "Blitzer", 100, 7, 3, 4, 8, ["Block"], "GA", "SP"],
    ["Dark Elf", 2, "Witch Elf", 110, 7, 3, 4, 7, ["Frenzy", "Dodge", "Jump Up"], "GA", "SP"],

    ["Dwarf", 16, "Blocker", 70, 4, 3, 2, 9, ["Block", "Tackle", "Thick Skull"], "GS", "AP"],
    ["Dwarf", 2, "Runner", 80, 6, 3, 3, 8, ["Sure Hands", "Thick Skull"], "GP", "AS"],
    ["Dwarf", 2, "Blitzer", 80, 5, 3, 3, 9, ["Block", "Thick Skull"], "GS", "AP"],
    ["Dwarf", 2, "Troll Slayer", 90, 5, 3, 2, 8, ["Block", "Dauntless", "Frenzy", "Thick Skull"], "GS", "AP"],
    ["Dwarf", 1, "Deathroller", 160, 4, 7, 1, 10, ["Loner", "Break Tackle", "Dirty Player", "Juggernaut", "Mighty Blow", "No Hands", "Secret Weapon", "Stand Firm"], "S", "GAP"],

    ["Elven Union", 16, "Lineman", 60, 6, 3, 4, 7, [], "GA", "SP"],
    ["Elven Union", 2, "Thrower", 70, 6, 3, 4, 7, ["Pass"], "GAP", "S"],
    ["Elven Union", 4, "Catcher", 100, 6, 3, 4, 7, ["Catch", "Nerves of Steel"], "GA", "SP"],
    ["Elven Union", 2, "Blitzer", 110, 7, 3, 4, 8, ["Block", "Side Step"], "GA", "SP"],

    ["Goblin", 16, "Goblin", 40, 6, 2, 3, 7, ["Dodge", "Right Stuff", "Stunty"], "A", "GSP"],
    ["Goblin", 1, "Doom Diver", 60, 6, 2, 3, 7, ["Right Stuff", "Stunty", "Swoop"], "A", "GSP"],
    ["Goblin", 1, "Bomma", 40, 6, 2, 3, 7, ["Bombardier", "Dodge", "Secret Weapon", "Stunty"], "A", "GSP"],
    ["Goblin", 1, "Looney", 40, 6, 2, 3, 7, ["Chainsaw", "Secret Weapon", "Stunty"], "A", "GSP"],
    ["Goblin", 1, "Fanatic", 70, 3, 7, 3, 7, ["Ball & Chain", "No Hands", "Secret Weapon", "Stunty"], "S", "GAP"],
    ["Goblin", 1, "Ooligan", 70, 6, 2, 3, 7, ["Disturbing Presence", "Dodge", "Fan Favorite", "Right Stuff", "Stunty"], "A", "GSP"],
    ["Goblin", 1, "Pogoer", 110, 7, 2, 3, 7, ["Dodge", "Leap", "Stunty", "Very Long Legs"], "A", "GSP"],
    ["Goblin", 2, "Troll", 110, 4, 5, 1, 9, ["Loner", "Always Hungry", "Mighty Blow", "Really Stupid", "Regeneration", "Throw Team-Mate"], "S", "GAP"],

    ["Halfling", 16, "Hopeful", 30, 5, 2, 3, 6, ["Dodge", "Right Stuff", "Stunty"], "A", "GSP"],
    ["Halfling", 2, "Catcher", 50, 5, 2, 3, 6, ["Catch", "Dodge", "Right Stuff", "Sprint", "Stunty"], "A", "GSP"],
    ["Halfling", 2, "Hefty", 50, 5, 2, 3, 7, ["Dodge", "Fend", "Stunty"], "AP", "GS"],
    ["Halfling", 2, "Treeman", 120, 2, 6, 1, 10, ["Mighty Blow", "Stand Firm", "Strong Arm", "Take Root", "Thick Skull", "Throw Team-Mate", "Timmm-ber!"], "S", "GAP"],

    ["High Elf", 16, "Lineman", 70, 6, 3, 4, 8, [], "GA", "SP"],
    ["High Elf", 2, "Thrower", 90, 6, 3, 4, 8, ["Pass", "Safe Throw"], "GAP", "S"],
    ["High Elf", 4, "Catcher", 90, 8, 3, 4, 7, ["Catch"], "GA", "SP"],
    ["High Elf", 2, "Blitzer", 100, 7, 3, 4, 8, ["Block"], "GA", "SP"],

    ["Human", 16, "Lineman", 50, 6, 3, 3, 8, [], "G", "ASP"],
    ["Human", 4, "Catcher", 60, 8, 2, 3, 7, ["Catch", "Dodge"], "GA", "SP"],
    ["Human", 2, "Thrower", 70, 6, 3, 3, 8, ["Sure Hands", "Pass"], "GP", "AS"],
    ["Human", 4, "Blitzer", 90, 7, 3, 3, 8, ["Block"], "GS", "AP"],
    ["Human", 1, "Ogre", 140, 5, 5, 2, 9, ["Loner", "Bone-head", "Mighty Blow", "Thick Skull", "Throw Team-Mate"], "S", "GAP"],

    ["Khemri Tomb Kings", 16, "Skeleton", 40, 5, 3, 2, 7, ["Regeneration", "Thick Skull"], "G", "ASP"],
    ["Khemri Tomb Kings", 2, "Thro-Ra", 70, 6, 3, 2, 7, ["Pass", "Regeneration", "Sure Hands"], "GP", "AS"],
    ["Khemri Tomb Kings", 2, "Blitz-Ra", 90, 6, 3, 2, 8, ["Block", "Regeneration"], "GS", "AP"],
    ["Khemri Tomb Kings", 4, "Tomb Guardian", 100, 4, 5, 1, 9, ["Decay", "Regeneration"], "S", "GAP"],

    ["Lizardman", 16, "Skink", 60, 8, 2, 3, 7, ["Dodge", "Stunty"], "A", "GSP"],
    ["Lizardman", 6, "Saurus", 80, 6, 4, 1, 9, [], "GS", "AP"],
    ["Lizardman", 1, "Kroxigor", 140, 6, 5, 1, 9, ["Loner", "Bone-head", "Mighty Blow", "Prehensile Tail", "Thick Skull"], "S", "GAP"],

    ["Necromantic Horror", 16, "Zombie", 40, 4, 3, 2, 8, ["Regeneration"], "G", "ASP"],
    ["Necromantic Horror", 2, "Ghoul Runner", 70, 7, 3, 3, 7, ["Dodge"], "GA", "SP"],
    ["Necromantic Horror", 2, "Wight Blitzer", 90, 6, 3, 3, 8, ["Block", "Regeneration"], "GS", "AP"],
    ["Necromantic Horror", 2, "Flesh Golem", 110, 4, 4, 2, 9, ["Regeneration", "Stand Firm", "Thick Skull"], "GS", "AP"],
    ["Necromantic Horror", 2, "Werewolf", 120, 8, 3, 3, 8, ["Claws", "Frenzy", "Regeneration"], "GA", "SP"],

    ["Norse", 16, "Lineman", 50, 6, 3, 3, 7, ["Block"], "G", "ASP"],
    ["Norse", 2, "Thrower", 70, 6, 3, 3, 7, ["Block", "Pass"], "GP", "AS"],
    ["Norse", 2, "Catcher", 90, 7, 3, 3, 7, ["Block", "Dauntless"], "GA", "SP"],
    ["Norse", 2, "Berserker", 90, 6, 3, 3, 7, ["Block", "Frenzy", "Jump Up"], "GS", "AP"],
    ["Norse", 2, "Ulfwerenar", 110, 6, 4, 2, 8, ["Frenzy"], "GS", "AP"],
    ["Norse", 1, "Yhetee", 140, 5, 5, 1, 8, ["Loner", "Claws", "Disturbing Presence", "Frenzy", "Wild Animal"], "S", "GAP"],

    ["Nurgle", 16, "Rotter", 40, 5, 3, 3, 8, ["Decay", "Nurgle's Rot"], "GM", "ASP"],
    ["Nurgle", 4, "Pestigor", 80, 6, 3, 3, 8, ["Horns", "Nurgle's Rot", "Regeneration"], "GSM", "AP"],
    ["Nurgle", 4, "Bloater", 110, 4, 4, 2, 9, ["Disturbing Presence", "Foul Appearance", "Nurgle's Rot", "Regeneration"], "GSM", "AP"],
    ["Nurgle", 1, "Rotspawn", 140, 4, 5, 1, 9, ["Loner", "Disturbing Presence", "Foul Appearance", "Mighty Blow", "Nurgle's Rot", "Really Stupid", "Regeneration", "Tentacles"], "S", "GAPM"],

    ["Ogre", 16, "Runt", 20, 5, 1, 3, 5, ["Dodge", "Right Stuff", "Side Step", "Stunty", "Titchy"], "A", "GSP"],
    ["Ogre", 6, "Ogre", 140, 5, 5, 2, 9, ["Bone-head", "Mighty Blow", "Thick Skull", "Throw Team-Mate"], "S", "GAP"],

    ["Orc", 16, "Lineman", 50, 5, 3, 3, 9, [], "G", "ASP"],
    ["Orc", 4, "Goblin", 40, 6, 2, 3, 7, ["Right Stuff", "Dodge", "Stunty"], "A", "GSP"],
    ["Orc", 2, "Thrower", 70, 5, 3, 3, 8, ["Sure Hands", "Pass"], "GP", "AS"],
    ["Orc", 4, "Black Orc", 80, 4, 4, 2, 9, [], "GS", "AP"],
    ["Orc", 4, "Blitzer", 80, 6, 3, 3, 9, ["Block"], "GS", "AP"],
    ["Orc", 1, "Troll", 110, 4, 5, 1, 9, ["Loner", "Always Hungry", "Mighty Blow", "Really Stupid", "Regeneration", "Throw Team-Mate"], "S", "GAP"],

    ["Skaven", 16, "Lineman", 50, 7, 3, 3, 7, [], "G", "ASPM"],
    ["Skaven", 2, "Thrower", 70, 7, 3, 3, 7, ["Pass", "Sure Hands"], "GP", "ASM"],
    ["Skaven", 4, "Gutter Runner", 80, 9, 2, 4, 7, ["Dodge", "Weeping Dagger"], "GA", "SPM"],
    ["Skaven", 2, "Blitzer", 90, 4, 4, 2, 9, ["Block"], "GS", "APM"],
    ["Skaven", 1, "Rat Ogre", 150, 6, 3, 3, 9, ["Loner", "Frenzy", "Mighty Blow", "Prehensile Tail", "Wild Animal"], "S", "GAPM"],

    ["Shambling Undead", 16, "Skeleton", 40, 5, 3, 2, 7, ["Regeneration", "Thick Skull"], "G", "ASP"],
    ["Shambling Undead", 16, "Zombie", 40, 4, 3, 2, 8, ["Regeneration"], "G", "ASP"],
    ["Shambling Undead", 4, "Ghoul Runner", 70, 7, 3, 3, 7, ["Dodge"], "GA", "SP"],
    ["Shambling Undead", 2, "Wight Blitzer", 90, 6, 3, 3, 8, ["Block", "Regeneration"], "GS", "AP"],
    ["Shambling Undead", 2, "Mummy", 120, 3, 5, 1, 9, ["Mighty Blow", "Regeneration"], "S", "GAP"],

    ["Vampire", 16, "Thrall", 40, 6, 3, 3, 7, [], "G", "ASP"],
    ["Vampire", 6, "Vampire", 110, 6, 4, 4, 8, ["Blood Lust", "Hypnotic Gaze", "Regeneration"], "GAS", "P"],

    ["Wood Elf", 16, "Lineman", 70, 7, 3, 4, 7, [], "GA", "SP"],
    ["Wood Elf", 4, "Catcher", 90, 8, 2, 4, 7, ["Catch", "Dodge", "Sprint"], "GA", "SP"],
    ["Wood Elf", 2, "Thrower", 90, 7, 3, 4, 7, ["Pass"], "GAP", "S"],
    ["Wood Elf", 2, "Wardancer", 120, 8, 3, 4, 7, ["Block", "Dodge", "Leap"], "GA", "SP"],
    ["Wood Elf", 1, "Treeman", 120, 2, 6, 1, 10, ["Loner", "Mighty Blow", "Stand Firm", "Strong Arm", "Take Root", "Thick Skull", "Throw Team-Mate"], "S", "GAP"],

    ["Bretonnian", 16, "Lineman", 40, 6, 3, 2, 7, ["Fend"], "G", "ASP"],
    ["Bretonnian", 4, "Yeoman", 70, 6, 3, 3, 8, ["Wrestle"], "GS", "AP"],
    ["Bretonnian", 4, "Blitzer", 110, 7, 3, 3, 8, ["Block", "Catch", "Dauntless"], "GAP", "S"],

    ["Daemons of Khorne", 16, "Pit Fighter", 60, 6, 3, 4, 8, ["Frenzy"], "GP", "AS"],
    ["Daemons of Khorne", 4, "Bloodletter", 80, 6, 3, 3, 7, ["Horns", "Juggernaut", "Regeneration"], "GAS", "P"],
    ["Daemons of Khorne", 2, "Herald", 90, 6, 3, 3, 8, ["Frenzy", "Horns", "Juggernaut"], "GS", "AP"],
    ["Daemons of Khorne", 1, "Bloodthirster", 180, 6, 5, 1, 9, ["Loner", "Frenzy", "Wild Animal", "Claws", "Horns", "Juggernaut", "Regeneration"], "S", "GAP"],

    ["Simyin", 12, "Bonobo", 50, 6, 3, 3, 7, ["Extra Arms"], "G", "ASP"],
    ["Simyin", 2, "Orangutan", 70, 5, 3, 3, 8, ["Extra Arms", "Strong Arm"], "GP", "AS"],
    ["Simyin", 2, "Chimpanzee", 80, 7, 3, 3, 7, ["Extra Arms", "Wrestle"], "GA", "SP"],
    ["Simyin", 4, "Gorilla", 100, 5, 4, 2, 8, ["Extra Arms", "Grab"], "AS", "GP"],
    ["Simyin", 1, "Silverback", 140, 5, 5, 1, 9, ["Loner", "Wild Animal", "Extra Arms", "Grab", "Mighty Blow"], "S", "GAP"],

    ["Chaos Pact", 12, "Marauder", 50, 6, 3, 3, 8, [], "GSPM", "A"],
    ["Chaos Pact", 1, "Goblin Renegade", 40, 6, 2, 3, 7, ["Animosity", "Dodge", "Right Stuff", "Stunty"], "AM", "GSP"],
    ["Chaos Pact", 1, "Skaven Renegade", 50, 7, 3, 3, 7, ["Animosity"], "GM", "ASP"],
    ["Chaos Pact", 1, "Dark Elf Renegade", 70, 6, 3, 4, 8, ["Animosity"], "GAM", "SP"],
    ["Chaos Pact", 1, "Orc Renegade", 50, 5, 3, 3, 9, ["Animosity"], "GM", "ASP"],
    ["Chaos Pact", 1, "Troll", 110, 4, 5, 1, 9, ["Loner", "Always Hungry", "Mighty Blow", "Really Stupid", "Regeneration", "Throw Team-Mate"], "S", "GAPM"],
    ["Chaos Pact", 1, "Ogre", 140, 5, 5, 2, 9, ["Loner", "Bone-head", "Mighty Blow", "Thick Skull", "Throw Team-Mate"], "S", "GAPM"],
    ["Chaos Pact", 1, "Minotaur", 150, 5, 5, 2, 8, ["Loner", "Frenzy", "Horns", "Mighty Blow", "Thick Skull", "Wild Animal"], "S", "GAPM"],

    ["Slann", 16, "Lineman", 60, 6, 3, 3, 8, ["Leap", "Very Long Legs"], "G", "ASP"],
    ["Slann", 4, "Catcher", 80, 7, 2, 4, 7, ["Diving Catch", "Leap", "Very Long Legs"], "GA", "SP"],
    ["Slann", 4, "Blitzer", 110, 7, 3, 3, 8, ["Diving Tackle", "Jump Up", "Leap", "Very Long Legs"], "GAS", "P"],
    ["Slann", 1, "Kroxigor", 140, 6, 5, 1, 9, ["Loner", "Bone-head", "Mighty Blow", "Prehensile Tail", "Thick Skull"], "S", "GAP"],

    ["Underworld Denizens", 12, "Goblin", 40, 6, 2, 3, 7, ["Right Stuff", "Dodge", "Stunty"], "AM", "GSP"],
    ["Underworld Denizens", 2, "Skaven Lineman", 50, 7, 3, 3, 7, ["Animosity"], "GM", "ASP"],
    ["Underworld Denizens", 2, "Skaven Thrower", 70, 7, 3, 3, 7, ["Animosity", "Pass", "Sure Hands"], "GPM", "AS"],
    ["Underworld Denizens", 2, "Skaven Blitzer", 90, 7, 3, 3, 8, ["Animosity", "Block"], "GSM", "AP"],
    ["Underworld Denizens", 1, "Troll", 110, 4, 5, 1, 9, ["Loner", "Always Hungry", "Mighty Blow", "Really Stupid", "Regeneration", "Throw Team-Mate"], "SM", "GAP"]
]
STAR_PLAYERS = [
    ["\"Rotten\" Rick Bupkeis", ["Necromantic Horror", "Shambling Undead"], ["Loner", "Dirty Player", "Regeneration", "Sneaky Git"], 110, 4, 3, 2, 8, ""],
    ["Asperon Thorn", ["Dark Elf", "Elven Union", "High Elf"], ["Loner", "Hail Mary Pass", "Kick-Off Return", "Pass", "Safe Throw", "Sure Hands"], 160, 6, 3, 4, 8, ""],
    ["Barik Farblast", ["Dwarf"], ["Loner", "Hail Mary Pass", "Pass", "Secret Weapon", "Strong Arm", "Sure Hands", "Thick Skull"], 60, 6, 3, 3, 8, ""],
    ["Bertha Bigfist", ["Amazon", "Halfling", "Ogre"], ["Loner", "Bone-head", "Break Tackle", "Dodge", "Mighty Blow", "Thick Skull", "Throw Team-Mate"], 290, 6, 5, 2, 9, ""],
    ["Big Jobo Hairyfoot", ["Halfling", "Ogre"], ["Loner", "Dirty Player", "Stand Firm", "Stunty", "Tackle", "Wrestle"], 120, 4, 3, 2, 8, ""],
    ["Bilerot Vomitflesh", ["Chaos Chosen", "Nurgle"], ["Loner", "Dirty Player", "Disturbing Presence", "Foul Appearance"], 180, 4, 5, 2, 9, ""],
    ["Bo Gallante", ["High Elf"], ["Loner", "Dodge", "Side Step", "Sprint", "Sure Feet"], 160, 8, 3, 4, 7, ""],
    ["Bomber Dribblesnot", ["Goblin", "Ogre", "Orc", "Chaos Pact", "Underworld"], ["Loner", "Accurate", "Bombardier", "Dodge", "Right Stuff", "Secret Weapon", "Stunty"], 60, 6, 2, 3, 7, ""],
    ["Boomer Eziasson", ["Dwarf", "Norse"], ["Loner", "Accurate", "Block", "Bombardier", "Secret Weapon", "Thick Skull"], 60, 6, 2, 3, 9, ""],
    ["Brick Far'th", ["Chaos", "Nurgle", "Ogre"], ["Loner", "Bone-head", "Mighty Blow", "Nerves of Steel", "Strong Arm", "Thick Skull", "Throw Team-Mate"], 290, 5, 5, 2, 9, "Note: When inducing this player, you must have 2 slots open on your roster so that you can also induce Grotty. This counts only as one Star Player inducement."],
    ["Bryce \"The Slice\" Cambuel", ["Khemri Tomb Kings", "Shambling Undead"], ["Loner", "Chainsaw", "Regeneration", "Secret Weapon", "Stand Firm", "Thick Skull"], 130, 5, 3, 2, 8, ""],
    ["Bulla Shardhorn", ["Nurgle"], ["Loner", "Block", "Extra Arms", "Foul Appearance", "Horns", "Nurgle's Rot", "Regeneration", "Stab", "Two Heads"], 230, 6, 3, 3, 8, ""],
    ["Captain Colander", ["Halfling"], ["Loner", "Catch", "Disturbing Presence", "Dodge", "Fend", "Jump Up", "Regeneration", "Right Stuff", "Side Step", "Stunty"], 100, 6, 2, 3, 7, ""],
    ["Cindy Piewhistle", ["Halfling"], ["Loner", "Accurate", "Bombardier", "Dodge", "Secret Weapon", "Stunty"], 50, 5, 2, 3, 6, ""],
    ["Count Luthor Von Drakenborg", ["Necromantic", "Undead", "Vampire"], ["Loner", "Block", "Hypnotic Gaze", "Regeneration", "Side Step"], 390, 6, 5, 4, 9, ""],
    ["Crazy Igor", ["Vampire", "Chaos Pact"], ["Loner", "Dauntless", "Regeneration", "Thick Skull"], 120, 6, 3, 3, 8, "Note: Crazy Igor can be bitten by a Vampire on your team as if he was a Thrall."],
    ["Curnoth Darkwold", ["Wood Elf"], ["Loner", "Dodge", "Frenzy", "Jump Up", "Leap", "Wrestle"], 240, 7, 3, 4, 7, ""],
    ["Deeproot Strongbranch", ["Halfling"], ["Loner", "Block", "Mighty Blow", "Stand Firm", "Strong Arm", "Thick Skull", "Throw Team-Mate", "Timmm-ber!"], 300, 2, 7, 1, 10, ""],
    ["Dolfar Longstride", ["Elf", "High Elf", "Wood Elf", "Bretonnian"], ["Loner", "Diving Catch", "Hail Mary Pass", "Kick", "Kick-Off Return", "Pass Block"], 150, 7, 3, 4, 7, ""],
    ["Eldril Sidewinder", ["Dark Elf", "Elf", "High Elf", "Wood Elf"], ["Loner", "Catch", "Dodge", "Hypnotic Gaze", "Nerves of Steel", "Pass Block"], 200, 8, 3, 4, 7, ""],
    ["Elijah Doom", ["Dark Elf", "Elven Union"], ["Loner", "Fend", "Guard", "Stand Firm", "Wrestle"], 190, 6, 3, 4, 9, ""],
    ["Fezglitch", ["Skaven", "Underworld"], ["Loner", "Ball & Chain", "Disturbing Presence", "Foul Appearance", "No Hands", "Secret Weapon"], 100, 4, 7, 3, 7, ""],
    ["Flint Churnblade", ["Dwarf"], ["Loner", "Block", "Chainsaw", "Secret Weapon", "Thick Skull"], 130, 5, 3, 2, 8, ""],
    ["Frank N Stein", ["Human", "Necromantic Horror", "Shambling Undead"], ["Loner", "Break Tackle", "Mighty Blow", "Regeneration", "Stand Firm", "Thick Skull"], 270, 4, 5, 1, 9, ""],
    ["Fungus the Loon", ["Goblin"], ["Loner", "Ball & Chain", "Mighty Blow", "No Hands", "Secret Weapon", "Stunty"], 80, 4, 7, 3, 7, ""],
    ["Furious George", ["Simyin"], ["Loner", "Catch", "Dodge", "Extra Arms", "Pass Block", "Strip Ball", "Wrestle"], 170, 7, 3, 3, 7, ""],
    ["G'Ral Blodschuker", ["Necromantic Horror", "Shambling Undead", "Vampire"], ["Loner", "Catch", "Dodge", "Sure Feet", "Wrestle"], 160, 7, 3, 3, 7, ""],
    ["Glart Smashrip Jr.", ["Skaven", "Underworld"], ["Loner", "Block", "Claws", "Juggernaut"], 210, 7, 4, 3, 8, ""],
    ["Glart Smashrip Sr.", ["Skaven", "Underworld"], ["Loner", "Block", "Claws", "Grab", "Juggernaut", "Stand Firm"], 190, 5, 4, 2, 8, ""],
    ["Gloriel Summerbloom", ["Wood Elf"], ["Loner", "Accurate", "Dodge", "Pass", "Side Step", "Sure Hands"], 160, 7, 2, 4, 7, ""],
    ["Gobbler Grimlich", ["Chaos Chosen", "Chaos Pact", "Underworld Denizens"], ["Loner", "Big Hand", "Disturbing Presence", "Leap", "Monstrous Mouth", "Regeneration", "Tentacles", "Very Long Legs"], 230, 5, 4, 2, 9, ""],
    ["Grashnak Blackhoof", ["Chaos", "Chaos Dwarf", "Nurgle", "Daemons of Khorne"], ["Loner", "Frenzy", "Horns", "Mighty Blow", "Thick Skull"], 310, 6, 6, 2, 8, ""],
    ["Gretchen Wachter", ["Necromantic Horror", "Shambling Undead", "Vampire"], ["Loner", "Disturbing Presence", "Dodge", "Foul Appearance", "Jump Up", "No Hands", "Regeneration", "Shadowing", "Sidestep"], 280, 7, 3, 4, 8, ""],
    ["Griff Oberwald", ["Human", "Bretonnian"], ["Loner", "Block", "Dodge", "Fend", "Sprint", "Sure Feet"], 320, 7, 4, 4, 8, ""],
    ["Grim Ironjaw", ["Dwarf"], ["Loner", "Block", "Dauntless", "Frenzy", "Multiple Block", "Thick Skull"], 220, 5, 4, 3, 8, ""],
    ["Grotty", ["Chaos", "Nurgle", "Ogre"], ["Loner", "Dodge", "Right Stuff", "Stunty"], 0, 6, 2, 4, 7, "Note: When inducing this player, you must have 2 slots open on your roster so that you can also induce Brick Far'th. This counts only as one Star Player inducement."],
    ["Guffle Pusmaw", ["Chaos Chosen", "Chaos Pact", "Nurgle"], ["Loner", "Foul Appearance", "Monstrous Mouth", "Nurgle's Rot"], 210000, 5, 3, 4, 9, ""],
    ["Hack Enslash", ["Khemri", "Necromantic", "Undead"], ["Loner", "Chainsaw", "Regeneration", "Secret Weapon", "Side Step"], 120, 6, 3, 2, 7, ""],
    ["Hakflem Skuttlespike", ["Skaven"], ["Loner", "Dodge", "Extra Arms", "Prehensile Tail", "Two Heads"], 200, 9, 3, 4, 7, ""],
    ["Headsplitter", ["Skaven"], ["Loner", "Frenzy", "Mighty Blow", "Prehensile Tail"], 340, 6, 6, 3, 8, ""],
    ["Helmut Wulf", ["Amazon", "Chaos Pact", "Human", "Lizardman", "Norse", "Vampire", "Slann"], ["Loner", "Chainsaw", "Secret Weapon", "Stand Firm"], 110, 6, 3, 3, 8, ""],
    ["Hemlock", ["Lizardman", "Slann"], ["Loner", "Block", "Dodge", "Side Step", "Jump Up", "Stab", "Stunty"], 170, 8, 2, 3, 7, ""],
    ["Horkon Heartripper", ["Dark Elf"], ["Loner", "Dodge", "Leap", "Multiple Block", "Shadowing", "Stab"], 210, 7, 3, 4, 7, ""],
    ["Hthark the Unstoppable", ["Chaos Dwarf"], ["Loner", "Block", "Break Tackle", "Juggernaut", "Sprint", "Sure Feet", "Thick Skull"], 330, 6, 5, 2, 9, ""],
    ["Hubris Rakarth", ["Dark Elf", "Elf"], ["Loner", "Block", "Dirty Player", "Jump Up", "Mighty Blow", "Strip Ball"], 260, 7, 4, 4, 8, ""],
    ["Humerus Carpal", ["Khemri"], ["Loner", "Catch", "Dodge", "Regeneration", "Nerves of Steel"], 130, 7, 2, 3, 7, ""],
    ["Icepelt Hammerblow", ["Norse"], ["Loner", "Claws", "Disturbing Presence", "Frenzy", "Regeneration", "Thick Skull"], 330, 5, 6, 1, 8, ""],
    ["Ithaca Benoin", ["Dark Elf", "Khemri"], ["Loner", "Accurate", "Dump-off", "Nerves of Steel", "Pass", "Regeneration", "Sure Hands"], 220, 7, 3, 3, 7, ""],
    ["Ivan \"The Animal\" Deathshroud", ["Khemri Tomb Kings", "Necromantic Horror", "Shambling Undead"], ["Loner", "Block", "Disturbing Presence", "Juggernaut", "Regeneration", "Strip Ball", "Tackle"], 230, 6, 4, 2, 8, ""],
    ["J Earlice", ["Necromantic", "Undead", "Vampire"], ["Loner", "Catch", "Diving Catch", "Dodge", "Sprint"], 180, 8, 3, 3, 7, ""],
    ["Jeremiah Kool", ["Dark Elf"], ["Loner", "Block", "Diving Catch", "Dodge", "Dump-off", "Kick-Off Return", "Nerves of Steel", "Pass", "Side Step"], 390, 8, 3, 5, 8, ""],
    ["Jordell Freshbreeze", ["Elf", "Wood Elf"], ["Loner", "Block", "Diving Catch", "Dodge", "Leap", "Side Step"], 260, 8, 3, 5, 7, ""],
    ["Karla Von Kill", ["Amazon", "Bretonnian", "Halfling", "Human", "Norse"], ["Loner", "Block", "Dauntless", "Dodge", "Jump Up"], 220, 6, 4, 3, 8, ""],
    ["King Boombastic", ["Simyin"], ["Loner", "Break Tackle", "Grab", "Piling On", "Wild Animal"], 270, 6, 5, 1, 9, ""],
    ["Kiroth Krakeneye", ["Dark Elf", "Elven Union"], ["Loner", "Disturbing Presence", "Foul Appearance", "Pass Block", "Tackle", "Tentacles"], 170, 7, 3, 4, 8, ""],
    ["Kreek Rustgouger", ["Skaven", "Underworld Denizens"], ["Loner", "Ball & Chain", "Mighty Blow", "No Hands", "Prehensile Tail", "Secret Weapon"], 130000, 5, 7, 2, 9, ""],
    ["Lewdgrip Whiparm", ["Chaos", "Nurgle", "Chaos Pact"], ["Loner", "Dodge", "Pass", "Strong Arm", "Sure Hands", "Tentacles"], 150, 6, 3, 3, 9, ""],
    ["Lord Borak the Despoiler", ["Chaos", "Nurgle"], ["Loner", "Block", "Dirty Player", "Mighty Blow"], 300, 5, 5, 3, 9, ""],
    ["Lottabottol", ["Lizardman", "Slann"], ["Loner", "Catch", "Diving Tackle", "Jump Up", "Leap", "Pass Block", "Shadowing", "Very Long Legs"], 220, 8, 3, 3, 8, ""],
    ["Lucien the Swift", ["Elf", "High Elf", "Wood Elf"], ["Loner", "Block", "Mighty Blow", "Tackle"], 390, 7, 3, 4, 8, "Note: When inducing this player, you must have 2 slots open on your roster so that you can also induce Valen the Swift. This counts only as one Star Player inducement."],
    ["Madcap Miggz", ["Goblin", "Underworld"], ["Loner", "Break Tackle", "Claws", "Leap", "Very Long Legs", "Wild Animal", "No Hands"], 170, 6, 4, 3, 8, ""],
    ["Maple Highgrove", ["Wood Elf"], ["Loner", "Grab", "Mighty Blow", "Stand Firm", "Tentacles", "Thick Skull"], 300, 3, 5, 1, 10, ""],
    ["Max Spleenripper", ["Chaos", "Nurgle"], ["Loner", "Chainsaw", "Secret Weapon"], 130, 5, 4, 3, 8, ""],
    ["Mighty Zug", ["Human", "Bretonnian"], ["Loner", "Block", "Mighty Blow"], 260, 4, 5, 2, 9, ""],
    ["Mordrix Hex", ["Dark Elf"], ["Loner", "Block", "Dauntless", "Dodge", "Fend", "Frenzy", "Mighty Blow"], 230, 7, 3, 4, 7, ""],
    ["Morg 'n' Thorg", ["All Except", "Khemri", "Necromantic", "Undead"], ["Loner", "Block", "Mighty Blow", "Thick Skull", "Throw Team-Mate"], 430, 6, 6, 3, 10, ""],
    ["Neddley Verruca", ["Halfling"], ["Loner", "Dodge", "Secret Weapon", "Stab", "Stunty", "Leap", "Very Long Legs"], 70, 5, 2, 3, 6, ""],
    ["Nobbla Blackwart", ["Chaos Dwarf", "Goblin", "Ogre", "Underworld"], ["Loner", "Block", "Dodge", "Chainsaw", "Secret Weapon", "Stunty"], 130, 6, 2, 3, 7, ""],
    ["Prince Moranion", ["Elf", "High Elf"], ["Loner", "Block", "Dauntless", "Tackle", "Wrestle"], 230, 7, 4, 4, 8, ""],
    ["Puggy Baconbreath", ["Halfling", "Human"], ["Loner", "Block", "Dodge", "Nerves of Steel", "Right Stuff", "Stunty"], 140, 5, 3, 3, 6, ""],
    ["Quetzal Leap", ["Lizardman", "Slann"], ["Loner", "Catch", "Diving Catch", "Fend", "Kick-Off Return", "Leap", "Nerves of Steel", "Very Long Legs"], 250, 8, 2, 4, 7, ""],
    ["Ramtut III", ["Khemri", "Necromantic", "Undead"], ["Loner", "Break Tackle", "Mighty Blow", "Regeneration", "Wrestle"], 380, 5, 6, 1, 7, ""],
    ["Rashnak Backstabber", ["Chaos Dwarf"], ["Loner", "Dodge", "Side Step", "Sneaky Git", "Stab"], 200, 7, 3, 3, 7, ""],
    ["Ripper Bolgrot", ["Goblin", "Orc"], ["Loner", "Grab", "Mighty Blow", "Regeneration", "Throw Team-Mate"], 270, 4, 6, 1, 9, ""],
    ["Roxanna Darknail", ["Amazon", "Dark Elf"], ["Loner", "Dodge", "Frenzy", "Jump Up", "Juggernaut", "Leap"], 250, 8, 3, 5, 7, ""],
    ["Rumbelow Sheepskin", ["Halfling"], ["Loner", "Block", "Horns", "Juggernaut", "No Hands", "Tackle", "Thick Skull"], 170, 6, 3, 3, 7, ""],
    ["Scrappa Sorehead", ["Goblin", "Ogre", "Orc"], ["Loner", "Dirty Player", "Dodge", "Leap", "Right Stuff", "Sprint", "Stunty", "Sure Feet", "Very Long Legs"], 150, 7, 4, 1, 9, ""],
    ["Scyla Anfingrimm", ["Chaos Chosen", "Norse"], ["Loner", "Claws", "Frenzy", "Prehensile Tail", "Thick Skull", "Wild Animal"], 250, 5, 5, 1, 9, ""],
    ["Setekh", ["Khemri", "Necromantic", "Undead"], ["Loner", "Block", "Break Tackle", "Juggernaut", "Regeneration", "Strip Ball"], 220, 6, 4, 2, 8, ""],
    ["Sinnedbad", ["Khemri", "Undead"], ["Loner", "Block", "Jump Up", "Pass Block", "Regeneration", "Secret Weapon", "Side Step", "Stab"], 80, 6, 3, 2, 7, ""],
    ["Skitter Stab-Stab", ["Skaven", "Underworld"], ["Loner", "Dodge", "Prehensile Tail", "Shadowing", "Stab"], 160, 9, 2, 4, 7, ""],
    ["Skrull Halfheight", ["Khemri Tomb Kings", "Shambling Undead"], ["Loner", "Accurate", "Nerves of Steel", "Pass", "Regeneration", "Sure Hands", "Thick Skull"], 190, 6, 3, 3, 8, ""],
    ["Slibli", ["Lizardman", "Slann"], ["Loner", "Block", "Grab", "Guard", "Stand Firm"], 250, 7, 4, 1, 8, ""],
    ["Soaren Hightower", ["High Elf"], ["Loner", "Fend", "Kick-Off Return", "Pass", "Safe Throw", "Sure Hands", "Strong Arm"], 180, 6, 3, 4, 8, ""],
    ["Swiftvine Glimmershard", ["Wood Elf"], ["Loner", "Disturbing Presence", "Fend", "Side Step", "Stab", "Stunty"], 130, 7, 2, 3, 6, ""],
    ["Throttlesnot \"The Impaler\"", ["Necromantic Horror", "Shambling Undead"], ["Loner", "Dirty Player", "Dodge", "Leap", "Regeneration", "Secret Weapon", "Stab", "Stunty"], 100, 6, 2, 3, 7, ""],
    ["Tolly Glocklinger", ["Nurgle"], ["Loner", "Ball & Chain", "Disturbing Presence", "Foul Appearance", "No Hands", "Nurgle's Rot", "Secret Weapon", "Stand Firm"], 110, 3, 7, 2, 9, ""],
    ["Ugroth Bolgrot", ["Orc", "Chaos Pact"], ["Loner", "Chainsaw", "Secret Weapon"], 100, 5, 3, 3, 9, ""],
    ["Valen the Swift", ["Elf", "High Elf", "Wood Elf"], ["Loner", "Accurate", "Nerves of Steel", "Pass", "Safe Throw", "Sure Hands"], 390, 7, 3, 5, 7, "Note: When inducing this player, you must have 2 slots open on your roster so that you can also induce Lucien the Swift. This counts only as one Star Player inducement."],
    ["Varag Ghoul-Chewer", ["Orc"], ["Loner", "Block", "Jump Up", "Mighty Blow", "Thick Skull"], 290, 6, 4, 3, 9, ""],
    ["Wilhelm Chaney", ["Necromantic", "Norse", "Vampire"], ["Loner", "Catch", "Claws", "Frenzy", "Regeneration", "Wrestle"], 240, 8, 4, 3, 8, ""],
    ["Willow Rosebark", ["Amazon", "Halfling", "Wood Elf", "Bretonnian"], ["Loner", "Dauntless", "Side Step", "Thick Skull"], 150, 5, 4, 3, 8, ""],
    ["Withergrasp Doubledrool", ["Chaos Chosen", "Chaos Pact", "Nurgle"], ["Loner", "Prehensile Tail", "Tackle", "Tentacles", "Two Heads", "Wrestle"], 170, 6, 3, 3, 8, ""],
    ["Zara the Slayer", ["Amazon", "Dwarf", "Halfling", "High Elf", "Human", "Norse", "Wood Elf", "Bretonnian"], ["Loner", "Block", "Dauntless", "Dodge", "Jump Up", "Stab", "Stakes"], 270, 6, 4, 3, 8, ""],
    ["Zolcath the Zoat", ["Amazon", "Lizardman", "Wood Elf"], ["Loner", "Disturbing Presence", "Juggernaut", "Mighty Blow", "Prehensile Tail", "Regeneration", "Sure Feet"], 280, 5, 5, 2, 9, ""],
    ["Zzharg Madeye", ["Chaos Dwarf", "Chaos Pact"], ["Loner", "Hail Mary Pass", "Pass", "Secret Weapon", "Strong Arm", "Sure Hands", "Tackle", "Thick Skull"], 90, 4, 4, 2, 9, ""]
]

INDUCEMENTS = [
    ["Bloodweiser Babe", 2, {"All": 50}, "The team purchases a keg of extra-special Bloodweiser magic ale for 50,000 gold pieces, and gets a lovely lady to serve players before going out for each drive. The combination of the ale and the young lady serving it means that for each purchase of this inducement, players on the team gain a +1 modifier to recover from KO'd for this match."],
    ["Bribe", 3, {"Goblin": 50, "Otherwise": 100}, "Goblin teams may buy a bribe for 50,000 gold pieces; any other team can buy a bribe for 100,000 gold pieces. Each bribe allows a team to attempt to ignore one call by the referee for a player who has committed a foul to be sent off, or a player armed with a secret weapon to be banned from the match. Roll a D6: on a roll of 2-6 the bribe is effective (preventing a turnover if the player was ejected for fouling), but on a roll of 1 the bribe is wasted and the call still stands! Each bribe may be used once per match."],
    ["Extra Team Training", 4, {"All": 100}, "Each extra team training session costs 100,000 gold pieces and allows the team to take one extra Team re-roll that may be used for this match only."],
    ["Halfling Master Chef", 1, {"Halfling": 100, "Otherwise": 300}, "Halfling teams may hire a Halfling Master Chef for 100,000 gold pieces; any other team can hire the Chef for 300,000 gold pieces. Roll 3D6 at the start of each half to see what effect the chef's cooking has on the team. For each dice that rolls 4 or more, the team is so inspired that they gain a Team Re-roll, and in addition, the opposing team is so distracted by the fantastic cooking smells emanating from their opponent's dug-out that they lose a Team Re-roll (but only if they have any left to lose)."],
    ["Igor", 1, {"All": 100}, "Any team that cannot purchase a permanent Apothecary can hire an Igor for 100,000 gold pieces to assist the team. An Igor is a master of needle and thread on rotting flesh, connecting hip bone to leg bone, rewrapping funeral wraps and so on. He can really get the boys shambling back to the pitch. An Igor may only be used once per a game to re-roll one failed Regeneration roll for a player."],
    ["Wandering Apothecary", 2, {"All": 100}, "Any team may hire a Wandering Apothecary or two to help during the match for 100,000 gold pieces each, if the team can normally purchase a permanent Apothecary. Often these Apothecaries are powerful priests of the local deity. While they would never allow themselves to be a permanent part of a heathen Blood Bowl team, they have been known to assist for a single match for a generous donation to their faith. The rules for Wandering Apothecaries are identical to the rules for purchased Apothecaries on page 17. Only one Apothecary may be used to re-roll each Casualty roll."],
    ["Wizard", 1, {"All": 150},
     "A team may hire a Wizard to help your team during the match for 150,000 gold pieces. Wizards, just like everybody else in the Old World, are keen sports fans and many are fanatically loyal in support of their chosen team. It is not surprising, then, that soon after the game was born, Wizards started 'helping out' the team they supported with carefully selected spells. Soon games were awash with magic as rival Wizards battled to give their team the edge. In the end, the Colleges of Magic were forced to insist that only teams that had bought a special license from the Colleges of Magic were allowed to have magical assistance. They limited this assistance to one spell per match, and even this had to be chosen from a very limited selection and cast by an officially appointed Colleges of Magic team Wizard. Wizards and fans alike soon realised that they really wanted to see a proper Blood Bowl match rather than a spellcasting contest, so the new rules were soon universally accepted.\nAny team is allowed to hire a Wizard for a match, as long as they can afford the whopping licensing fee charged by the College of Magic concerned. No team may hire more than one Wizard per match. Wizards can be represented in games with one of the Wizard models from the range of Citadel miniatures for Warhammer. This isn't strictly necessary, but looks a lot better than representing a Wizard with a bottle top or tiddlywink!\nOnce per game, the Wizard is allowed to cast either a Fireball spell or a Lightning Bolt spell. Wizards may only cast spells at the start of his own turn before any player performs an Action or immediately after his own team's turn has ended even if it ended with a turnover.\nFireball: Choose a target square anywhere on the pitch. Roll one dice to hit each standing player (from either team) that is either in the target square or a square adjacent to it. If the 'to hit' roll is a 4 or more then the target is Knocked Down. If it is a 3 or less he manages to dodge the fireball's blast. Make an Armour roll (and possible Injury as well) for any player that is Knocked Down as if they had been Knocked Down by a player with the Mighty Blow skill. If a player on the moving team is Knocked Down by a fireball, then the moving team does not suffer a turnover unless the player was carrying the ball at the time.\nLightning Bolt: Pick a standing player anywhere on the pitch, and roll one dice. If the score is a 2 or higher, then he has been hit by the lightning bolt. If the roll is a 1 then he manages to dodge out of the way. A player hit by a lightning bolt is Knocked Down and must make an Armour roll (and possible Injury as well) as if hit by a player with the Mighty Blow skill."],
    ["Cheerleader", 3, {"All": 20}, "Most Blood Bowl teams have a troupe or two of cheerleaders both to inspire the team's players and their fans. It's the team's cheerleaders' job to whip the fans into a state of frenzy and lead the chanting and singing as the crowd's shouts and howls build up to a deafening crescendo. The more cheerleaders you have on your team, the more likely you are to win the 'Cheering Fans' result on the Kick-Off table"],
    ["Assistant Coach", 3, {"All": 20}, "Assistant coaches include offensive and defensive coordinators, special team coaches, personal trainers for your legendary players and numerous others. As a team becomes more successful the number of assistant coaches on its roster just seems to grow and grow. The more assistant coaches you have on your team, the more likely you are to win the 'Brilliant Coaching' result on the Kick-Off table"],
    ["Marketing Blitz", 5, {"All": 20}, "+1 Fan Factor"]
]

PURCHASES = [
    ["Cheerleader", None, {"All": 10}, "Most Blood Bowl teams have a troupe or two of cheerleaders both to inspire the team's players and their fans. It's the team's cheerleaders' job to whip the fans into a state of frenzy and lead the chanting and singing as the crowd's shouts and howls build up to a deafening crescendo. The more cheerleaders you have on your team, the more likely you are to win the 'Cheering Fans' result on the Kick-Off table"],
    ["Assistant Coach", None, {"All": 10}, "Assistant coaches include offensive and defensive coordinators, special team coaches, personal trainers for your legendary players and numerous others. As a team becomes more successful the number of assistant coaches on its roster just seems to grow and grow. The more assistant coaches you have on your team, the more likely you are to win the 'Brilliant Coaching' result on the Kick-Off table"],
    ["Apothecary", 1, {"All": 50, "Exceptions": ["Khemri", "Necromantic", "Nurgle", "Undead"]},
     "An Apothecary is a healer wise in the ways of medicine and the healing arts who looks after the injured players in a Blood Bowl team - and so has a strenuous and full-time job! It costs 50,000 gold pieces to purchase an Apothecary to permanently look after your team during a match. He may be represented by an appropriate Citadel miniature if you wish. A team may not have more than one purchased Apothecary. Khemri, Necromantic, Nurgle and Undead teams may not purchase or use an Apothecary.\nDuring a match, an Apothecary may attempt to cure a player who has suffered a Casualty or been KO'd. An Apothecary can be used only once per match. If the player was KO'd leave him on the pitch Stunned, or, if he was not on the pitch, put him in the Reserves box. Otherwise immediately after the player suffers the Casualty, you can use the Apothecary to make your opponent roll again on the Casualty table (see page 25) and then you choose which of the two results to apply. If the player is only Badly Hurt after this roll (even if it was the original Casualty roll) the Apothecary has managed to patch him up and pump him full of painkillers so that the player may be moved into the Reserves box."],
    ["Team Re-roll", 8, {"Amazon": 50, "Chaos": 60, "Chaos Dwarf": 70, "Dark Elf": 50, "Dwarf": 50, "Elf": 50, "Goblin": 60, "Halfling": 60, "High Elf": 50, "Human": 50, "Khemri": 70, "Lizardman": 60, "Necromantic": 70, "Norse": 60, "Nurgle": 70, "Ogre": 70, "Orc": 60, "Skaven": 60, "Undead": 70, "Vampire": 70, "Wood Elf": 50, "Chaos Pact": 70, "Slann": 50, "Underworld": 70, "Bretonnian": 70, "Daemons of Khorne": 70, "Simyin": 60}, "Team re-rolls represent how well trained a team is. A coach may use a team re-roll to re-roll any dice roll (other than Scatter, Distance, Direction, Armour, Injury or Casualty rolls) made by a player on his own team and who is still on the pitch during their own turn (even if the dice roll was successful). The result of the new roll must be accepted in place of the first, even if it is worse. A coach may not use more than one Re-roll counter per turn, and may not use a Re-roll counter to force the opposing coach to re-roll a dice roll."],
    ["Bribe", 2, {"Goblin": 100, "Otherwise": 200}, "Goblin teams may roster a bribe for 100,000 gold pieces; any other team can buy a bribe for 200,000 gold pieces. Each bribe allows a team to attempt to ignore one call by the referee for a player who has committed a foul to be sent off, or a player armed with a secret weapon to be banned from the match. Roll a D6: on a roll of 2-6 the bribe is effective (preventing a turnover if the player was ejected for fouling), but on a roll of 1 the bribe is wasted and the call still stands! Each bribe may be used once per match."],
    ["Halfling Master Chef", 1, {"Halfling": 200, "Otherwise": 600}, "Halfling teams may roster a Halfling Master Chef for 200,000 gold pieces; any other team can roster the Chef for 600,000 gold pieces. Roll 3D6 at the start of each half to see what effect the chef's cooking has on the team. For each dice that rolls 4 or more, the team is so inspired that they gain a Team Re-roll, and in addition, the opposing team is so distracted by the fantastic cooking smells emanating from their opponent's dug-out that they lose a Team Re-roll (but only if they have any left to lose)."]
]

MATCH_TYPES = ["Preseason", "Regular", "Postseason", "Championship"]

COACHES = {
    "Matt": {"email": "mattiequinn35@gmail.com", "aka": ["Matthew Quinn", "Matthew"]},
    "Troy": {"email": "troychrisman@gmail.com", "aka": ["TDC", "Troy Chrisman"]},
    "Phillip": {"email": "pdponzer@gmail.com", "aka": ["Phillip Ponzer", "Phil"]},
    "Bryan": {"email": "bmankc2001@gmail.com", "aka": ["Bryan Bledsoe"]},
    "Chris": {"email": "ckthomason@gmail.com", "aka": ["The Spirit of Harambe", "Chris Thomason", "Patches O'Houlihan"]},
    "Trever": {"email": "guidedbyhim@gmail.com", "aka": ["Trever Leikam", "Sun Tzu", "Trever \"Sun Tzu\" Leikam"]},
    "Daniel": {"email": "dwconnors84@gmail.com", "aka": ["Daniel Connors"]},
    "Joe R": {"email": "vanhalfling@gmail.com", "aka": ["Joe Roberts", "Joe \"profgoldfinch\" Roberts"]},
    "James V": {"email": "jdv311@gmail.com", "aka": ["James Valdez"]},
    "David": {"email": "godzfirefly@gmail.com", "aka": ["David J Harshman"]},
    "Kendal": {"email": "jon.kbowser@gmail.com", "aka": ["Kendal Bowser", "Kendal Jon Bowser"]},
    "Patrick": {"email": "opcubby@gmail.com", "aka": ["Patrick Morissey", "Patrick Morrissey"]},
    "Aaron M": {"email": "ammckinn90@gmail.com", "aka": ["Aaron Mckinney"]},
    "Flagg": {"email": "eyeofthedragon101@gmail.com", "aka": ["Flagg Thorin"]},
    "Randy": {"email": "rkrg.enterprises13@gmail.com", "aka": ["Randy Cummins"]},
    "Joe H": {"email": "hudgenslaw@gmail.com", "aka": ["Joe Hudgens"]},
    "Russell": {"email": "christiangeek6170@gmail.com", "aka": ["Russell Parks"]},
    "Nicholas": {"email": "moluccowrathe@gmail.com", "aka": ["Nicholas J Milberger", "Nicholas J. Milberger"]},
    "Luke": {"email":"lukermichaels@gmail.com", "aka": ["Luke Michaels"]},
    "Ron": {"email":"chevyman1ton@gmail.com", "aka": ["Ron Bowser"]}
}

TEAMS = [
    ["Sylvania Suckhawks", "Joe R", "Vampire", "Season 1", []],
    ["Jacksonville Thrillers", "Joe R", "Necromantic", "Season 1", []],
    ["Population: Orc", "Phillip", "Orc", "Season 1", []],
    ["Wrath of the Month", "Chris", "Amazon", "Season 1", []],
    ["OrCares", "Troy", "Orc", "Season 1", []],
    ["Guardians of the Nile", "Kendal", "Lizardman", "Season 1", []],
    ["Skyhaven Silverfoxes", "Chris", "Elf", "Season 1", []],
    ["Ashtown Villains", "Joe H", "Undead", "Season 1", []],
    ["Aedra & Daedra", "Kendal", "Dark Elf", "Season 1", []],
    ["Hellions of Troy", "Troy", "Human", "Season 1", []],
    ["Komodo Sinodons", "Randy", "Lizardman", "Season 2", []],
    ["Population: Orc", "Phillip", "Orc", "Season 2", []],
    ["Quixotic Crushers", "Matthew", "Bretonnian", "Season 2", []],
    ["Tatooine Technographers", "Joe R", "Goblin", "Season 2", []],
    ["Riverdale Ravagers", "Joe R", "Chaos Pact", "Season 2", []],
    ["Blood Mountain Berserkers", "David", "Dwarf", "Season 2", []],
    ["Bows and Bandannas", "Kendal", "Necromantic", "Season 2", []],
    ["Fink Foulers", "Daniel", "Skaven", "Season 2", []],
    ["Hellions of Troy", "Troy", "Human", "Season 2", []],
    ["Mile High Maulers", "Joe H", "Norse", "Season 2", []],
    ["The Replacements", "Trever", "Human", "Season 2", []],
    ["Turtle Isle Typhoons", "Chris", "Simyin", "Season 2", []],
    ["The Sleeping Giants", "Troy", "Undead", "Season 2", []]
]

DECKS = [
    ["Misc Mayhem", 50],
    ["Special Team Plays", 50],
    ["Magic Items", 50],
    ["Dirty Tricks", 50],
    ["Good Karma", 100],
    ["Random Events", 200],
    ["Desperate Measures", 400]
]
# 13+13+13+13+26+18+8 = 104
CARDS = [
    ["Badyear Git", "2H", "Misc Mayhem", "A goblin doom diver who was too cheap to pay for admission is hit by the kick-off while flying over the stadium.", "Play at any kick-off after all players have been set up and the ball placed, but before any scatter has been rolled.", "The ball scatters 2d6, instead of 1d6, on this kick-off."],
    ["Sprinkler Malfunction", "3H", "Misc Mayhem", "", "", ""],
    ["Eclipse", "4H", "Misc Mayhem", "", "", ""],
    ["Fanatic Invasion", "5H", "Misc Mayhem", "", "", ""],
    ["Friendly Fans", "6H", "Misc Mayhem", "", "", ""],
    ["Rowdy Fans", "7H", "Misc Mayhem", "", "", ""],
    ["Heckler", "8H", "Misc Mayhem", "", "", ""],
    ["Hometown Fans", "9H", "Misc Mayhem", "", "", ""],
    ["Incoming!", "10H", "Misc Mayhem", "", "", ""],
    ["Rogue Wizard", "JH", "Misc Mayhem", "", "", ""],
    ["Ball Clone", "QH", "Misc Mayhem", "", "", ""],
    ["Johnny Waterboy", "KH", "Misc Mayhem", "", "", ""],
    ["That Babe's Got Talent!", "AH", "Misc Mayhem", "", "", ""],
    ["Come To Papa!", "2D", "Special Team Plays", "", "", ""],
    ["Dogged Defense", "3D", "Special Team Plays", "", "", ""],
    ["Flea Flicker", "4D", "Special Team Plays", "", "", ""],
    ["Going the Extra Mile", "5D", "Special Team Plays", "", "", ""],
    ["Heroic Leap", "6D", "Special Team Plays", "", "", ""],
    ["New Blocking Scheme", "7D", "Special Team Plays", "", "", ""],
    ["Perfect Kick", "8D", "Special Team Plays", "", "", ""],
    ["Option Play", "9D", "Special Team Plays", "", "", ""],
    ["Punt", "10D", "Special Team Plays", "", "", ""],
    ["Spectacular Catch", "JD", "Special Team Plays", "", "", ""],
    ["Suicide Blitz", "QD", "Special Team Plays", "", "", ""],
    ["Wake Up Call", "KD", "Special Team Plays", "", "", ""],
    ["Beguiling Bracers", "AD", "Special Team Plays", "", "", ""],
    ["Belt of Invulnerability", "2C", "Magic Items", "", "", ""],
    ["Beguiling Bracers", "3C", "Magic Items", "", "", ""],
    ["Fawndough's Headband", "4C", "Magic Items", "", "", ""],
    ["Force Shield", "5C", "Magic Items", "", "", ""],
    ["Gikta's Strength of da Bear", "6C", "Magic Items", "", "", ""],
    ["Gloves of Holding", "7C", "Magic Items", "", "", ""],
    ["Inertia Dampner", "8C", "Magic Items", "", "", ""],
    ["Lucky Charm", "9C", "Magic Items", "", "", ""],
    ["Magic Gloves of Jark Longarm", "10C", "Magic Items", "", "", ""],
    ["Good Old Magic Codpiece", "JC", "Magic Items", "", "", ""],
    ["Rabbit's Foot", "QC", "Magic Items", "", "", ""],
    ["Ring of Teleportation", "KC", "Magic Items", "", "", ""],
    ["Wand of Smashing", "AC", "Magic Items", "", "", ""],
    ["Blatant Foul", "2S", "Dirty Tricks", "", "", ""],
    ["Chop Block", "3S", "Dirty Tricks", "", "", ""],
    ["Custard Pie", "4S", "Dirty Tricks", "", "", ""],
    ["Distract", "5S", "Dirty Tricks", "", "", ""],
    ["Greased Shoes", "6S", "Dirty Tricks", "", "", ""],
    ["Gromskull's Exploding Runes", "7S", "Dirty Tricks", "", "", ""],
    ["Illegal Substitution", "8S", "Dirty Tricks", "", "", ""],
    ["Kicking Boots", "9S", "Dirty Tricks", "", "", ""],
    ["Pit Trap", "10S", "Dirty Tricks", "", "", ""],
    ["Spiked Ball", "JS", "Dirty Tricks", "", "", ""],
    ["Stolen Playbook", "QS", "Dirty Tricks", "", "", ""],
    ["Trampoline Trap", "KS", "Dirty Tricks", "", "", ""],
    ["Witch's Brew", "AS", "Dirty Tricks", "", "", ""],
    ["All Out Blitz", "2H", "Good Karma", "", "", ""],
    ["Banana Skin", "3H", "Good Karma", "", "", ""],
    ["Butterfingers", "4H", "Good Karma", "", "", ""],
    ["Chainsaw", "5H", "Good Karma", "", "", ""],
    ["Dazed and Confused", "6H", "Good Karma", "", "", ""],
    ["Doc Bonesaw", "7H", "Good Karma", "", "", ""],
    ["Extra Training", "8H", "Good Karma", "", "", ""],
    ["Fan Uproar", "9H", "Good Karma", "", "", ""],
    ["Hurry Up Offense", "10H", "Good Karma", "", "", ""],
    ["Intensive Training", "JH", "Good Karma", "", "", ""],
    ["Unsportsmanlike Conduct", "QH", "Good Karma", "", "", ""],
    ["Knutt's Spell of Awesome Strength", "KH", "Good Karma", "", "", ""],
    ["Lewd Maneuvers", "AH", "Good Karma", "", "", ""],
    ["Lurve Potion", "2D", "Good Karma", "", "", ""],
    ["Magic Helmet", "3D", "Good Karma", "", "", ""],
    ["Miracle Worker", "4D", "Good Karma", "", "", ""],
    ["One with the Kicker", "5D", "Good Karma", "", "", ""],
    ["Razzle Dazzle", "6D", "Good Karma", "", "", ""],
    ["Suitable Pitch", "7D", "Good Karma", "", "", ""],
    ["Rune of Fear", "8D", "Good Karma", "", "", ""],
    ["Scutt's Scroll of Weather Magic", "9D", "Good Karma", "", "", ""],
    ["Stiletto", "10D", "Good Karma", "", "", ""],
    ["Team Anthem", "JD", "Good Karma", "", "", ""],
    ["The Fan", "QD", "Good Karma", "", "", ""],
    ["The Wall", "KD", "Good Karma", "", "", ""],
    ["Woof Woof!", "AD", "Good Karma", "", "", ""],
    ["Bad Habits", "2C", "Random Events", "", "", ""],
    ["Ballista", "3C", "Random Events", "", "", ""],
    ["Blackmail", "4C", "Random Events", "", "", ""],
    ["Buzzing", "5C", "Random Events", "", "", ""],
    ["Duh, Where am I?", "6C", "Random Events", "", "", ""],
    ["Ego Trip", "7C", "Random Events", "", "", ""],
    ["Zap!", "8C", "Random Events", "", "", ""],
    ["Gimme That!", "9C", "Random Events", "", "", ""],
    ["Iron Man", "10C", "Random Events", "", "", ""],
    ["Kid Gloves", "2S", "Random Events", "", "", ""],
    ["Knuckledusters", "3S", "Random Events", "", "", ""],
    ["Magic Sponge", "4S", "Random Events", "", "", ""],
    ["Mine", "5S", "Random Events", "", "", ""],
    ["Not-so-Secret Weapon", "6S", "Random Events", "", "", ""],
    ["Orcidas Sponsorship", "7S", "Random Events", "", "", ""],
    ["Rakarth's Curse of Petty Spite", "8S", "Random Events", "", "", ""],
    ["Tackling Machine", "9S", "Random Events", "", "", ""],
    ["Get 'em Lads!", "10S", "Random Events", "", "", ""],
    ["Assassin", "JC", "Desperate Measures", "", "", ""],
    ["Doom and Gloom", "QC", "Desperate Measures", "", "", ""],
    ["Da Freight Train", "KC", "Desperate Measures", "", "", ""],
    ["Morley's Revenge", "AC", "Desperate Measures", "", "", ""],
    ["I am the Greatest", "JS", "Desperate Measures", "", "", ""],
    ["Mindblow", "QS", "Desperate Measures", "", "", ""],
    ["Come on Boys!", "KS", "Desperate Measures", "", "", ""],
    ["Mysterious Old Medicine Man", "AS", "Desperate Measures", "", "", ""]
]


def getAliases(name, aliases):
    if name in aliases:
        return aliases[name]
    for alias, aliasList in aliases.iteritems():
        if alias.lower() == name.lower():
            return aliases[alias]
        for aliasName in aliasList:
            if aliasName.lower() == name.lower():
                return aliasList
    return []


if __name__ == '__main__':
    TEAMS = []
    t = Timer()
    t1 = Timer()
    totalRosters = 0
    totalMatches = 0
    if True:
        # Season 1 Rosters
        if True:
            print "Loading Season 1 Rosters..."
            s1 = load_workbook("imports/S1 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            # sheetNames = ["T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8"]  # , "T9", "T10"]
            sheetNames = s1.get_sheet_names()[5:13]
            for sheetName in sheetNames:
                rosterSheet = s1[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamInfo = rows[2]
                raceName = teamInfo[4]
                teamName = teamInfo[9]
                coachName = teamInfo[13]

                treasury = rows[38][12]
                teamValue = rows[39][17]
                rerolls = rows[33][12]
                fanFactor = rows[34][12]
                assistantCoaches = rows[35][12]
                cheerleaders = rows[36][12]
                apothecary = rows[37][12]

                players = []
                for row in rows[5:21]:
                    playerName = row[3]
                    playerPosition = row[4]
                    ma = row[6]
                    st = row[7]
                    ag = row[8]
                    av = row[9]
                    skillsTemp = row[10].split(",") if row[10] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    skillReplace = [
                        ["MA+", "+MA"],
                        ["ST+", "+ST"],
                        ["AG+", "+AG"],
                        ["AV+", "+AV"],
                        ["MA-", "-MA"],
                        ["ST-", "-ST"],
                        ["AG-", "-AG"],
                        ["AV-", "-AV"],
                    ]
                    for skillRepl in skillReplace:
                        skill1 = skillRepl[0]
                        skill2 = skillRepl[1]
                        if skill1 in skills:
                            skills.remove(skill1)
                            skills.append(skill2)

                    status = row[11]
                    completions = row[12]
                    touchdowns = row[13]
                    interceptions = row[14]
                    casualties = row[15]
                    kills = None
                    mvps = row[16]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                    # ["Sylvania Suckhawks", "Joe R", "Vampire", "Season 1", []],
                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 1", players, teamValue, 0, 0])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)

        # Season 2 Rosters
        if True:
            print
            print "Loading Season 2 Rosters..."
            t.start()
            s2 = load_workbook("imports/S2 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s2.get_sheet_names()[4:21]
            for sheetName in sheetNames:
                rosterSheet = s2[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[19][8]
                print "        %s" % teamName

                raceName = rows[20][8]
                coachName = rows[21][8]
                teamValue = rows[22][8]
                treasury = rows[23][8]
                rerolls = rows[19][19] if rows[19][19] is not None else rows[19][20]
                fanFactor = rows[20][19] if rows[20][19] is not None else rows[20][20]
                assistantCoaches = rows[21][19] if rows[21][19] is not None else rows[21][20]
                cheerleaders = rows[22][19] if rows[22][19] is not None else rows[22][20]
                apothecary = rows[23][19] if rows[23][19] is not None else rows[23][20]

                try:
                    cheerleaders = int(cheerleaders)
                except Exception as ex:
                    pass

                players = []
                for row in rows[2:18]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[13]
                    injuryST = row[14]
                    injuryAG = row[15]
                    injuryAV = row[16]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[18]
                    touchdowns = row[19]
                    interceptions = row[17]
                    casualties = row[20]
                    kills = row[21]
                    mvps = row[22]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 2", players, teamValue, 0, 0])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)

        # Season 3 Rosters
        if True:
            print
            print "Loading Season 3 Rosters..."
            t.start()
            s3 = load_workbook("imports/S3 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s3.get_sheet_names()[0:22]
            for sheetName in sheetNames:
                rosterSheet = s3[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[20][8]
                raceName = rows[21][8]
                coachName = rows[22][8]
                teamValue = rows[23][8]
                treasury = rows[24][8]
                rerolls = rows[20][19]
                fanFactor = rows[21][19]
                assistantCoaches = rows[22][19] if rows[22][19] else 0
                cheerleaders = rows[23][19] if rows[23][19] else 0
                apothecary = rows[24][19] if rows[24][19] else 0

                players = []
                for row in rows[3:19]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[15]
                    injuryST = row[16]
                    injuryAG = row[17]
                    injuryAV = row[18]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[20]
                    touchdowns = row[21]
                    interceptions = row[19]
                    casualties = row[22]
                    kills = row[23]
                    mvps = row[24]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 3", players, teamValue, 0, 0])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)

        # Season 3.5 Rosters
        if True:
            print
            print "Loading Season 3.5 Rosters..."
            t.start()
            s3p5 = load_workbook("imports/S3.5 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s3p5.get_sheet_names()[:27]
            for sheetName in sheetNames:
                rosterSheet = s3p5[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[20][8]
                raceName = rows[23][8]
                coachName = rows[24][8]
                teamValue = rows[25][8]
                treasury = rows[26][8]
                rerolls = rows[20][19]
                fanFactor = rows[21][19]
                assistantCoaches = rows[22][19] if rows[22][19] else 0
                cheerleaders = rows[23][19] if rows[23][19] else 0
                apothecary = rows[24][19] if rows[24][19] else 0
                masterChef = rows[25][19] if rows[25][19] else 0
                bribes = rows[26][19] if rows[26][19] else 0

                if rows[25][11] != "HALFLING MASTER CHEF": # Then this utilizes the old style sheet without master chef and bribes added
                    raceName = rows[21][8]
                    coachName = rows[22][8]
                    teamValue = rows[23][8]
                    treasury = rows[24][8]

                players = []
                for row in rows[3:19]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[15]
                    injuryST = row[16]
                    injuryAG = row[17]
                    injuryAV = row[18]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[20]
                    touchdowns = row[21]
                    interceptions = row[19]
                    casualties = row[22]
                    kills = row[23]
                    mvps = row[24]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 3.5", players, teamValue, masterChef, bribes])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)

        # Season 4 Rosters
        if True:
            print
            print "Loading Season 4 Rosters..."
            t.start()
            s4 = load_workbook("imports/S4 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s4.get_sheet_names()[:13]
            for sheetName in sheetNames:
                rosterSheet = s4[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[20][8]
                raceName = rows[23][8]
                coachName = rows[24][8]
                teamValue = rows[25][8]
                treasury = rows[26][8]
                rerolls = rows[20][19]
                fanFactor = rows[21][19]
                assistantCoaches = rows[22][19] if rows[22][19] else 0
                cheerleaders = rows[23][19] if rows[23][19] else 0
                apothecary = rows[24][19] if rows[24][19] else 0
                masterChef = rows[25][19] if rows[25][19] else 0
                bribes = rows[26][19] if rows[26][19] else 0

                if rows[25][11] != "HALFLING MASTER CHEF": # Then this utilizes the old style sheet without master chef and bribes added
                    raceName = rows[21][8]
                    coachName = rows[22][8]
                    teamValue = rows[23][8]
                    treasury = rows[24][8]

                players = []
                for row in rows[3:19]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[15]
                    injuryST = row[16]
                    injuryAG = row[17]
                    injuryAV = row[18]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[20]
                    touchdowns = row[21]
                    interceptions = row[19]
                    casualties = row[22]
                    kills = row[23]
                    mvps = row[25]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 4", players, teamValue, masterChef, bribes])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)
 
        # Season 5 Rosters
        if True:
            print
            print "Loading Season 5 Rosters..."
            t.start()
            s5 = load_workbook("imports/S5 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s5.get_sheet_names()[:13]
            for sheetName in sheetNames:
                rosterSheet = s5[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[20][8]
                raceName = rows[23][8]
                coachName = rows[24][8]
                teamValue = rows[25][8]
                treasury = rows[26][8]
                rerolls = rows[20][19]
                fanFactor = rows[21][19]
                assistantCoaches = rows[22][19] if rows[22][19] else 0
                cheerleaders = rows[23][19] if rows[23][19] else 0
                apothecary = rows[24][19] if rows[24][19] else 0
                masterChef = rows[25][19] if rows[25][19] else 0
                bribes = rows[26][19] if rows[26][19] else 0

                players = []
                for row in rows[3:19]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[15]
                    injuryST = row[16]
                    injuryAG = row[17]
                    injuryAV = row[18]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[20]
                    touchdowns = row[21]
                    interceptions = row[19]
                    casualties = row[22]
                    kills = row[23]
                    mvps = row[25]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 5", players, teamValue, masterChef, bribes])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)

        # Season 6 Rosters
        if True:
            print
            print "Loading Season 6 Rosters..."
            t.start()
            s6 = load_workbook("imports/S6 Rosters.xlsx", data_only=True, read_only=True)
            print "    Sheet loaded in %s! Now loading rosters..." % t.elapsed()
            t.start()
            sheetNames = s6.get_sheet_names()[:13]
            for sheetName in sheetNames:
                rosterSheet = s6[sheetName]
                rows = []
                for row in rosterSheet.rows:
                    rows.append([cell.value for cell in row])

                teamName = rows[20][8]
                raceName = rows[23][8]
                coachName = rows[24][8]
                teamValue = rows[25][8]
                treasury = rows[26][8]
                rerolls = rows[20][19]
                fanFactor = rows[21][19]
                assistantCoaches = rows[22][19] if rows[22][19] else 0
                cheerleaders = rows[23][19] if rows[23][19] else 0
                apothecary = rows[24][19] if rows[24][19] else 0
                masterChef = rows[25][19] if rows[25][19] else 0
                bribes = rows[26][19] if rows[26][19] else 0

                players = []
                for row in rows[3:19]:
                    playerName = row[2]
                    playerPosition = row[3]
                    ma = row[4]
                    st = row[5]
                    ag = row[6]
                    av = row[7]
                    skillsTemp = row[8].split(",") if row[8] else []
                    skills = [skill.strip() for skill in skillsTemp]
                    newSkills = row[9].split(",") if row[9] else []
                    skills.extend([newSkill.strip() for newSkill in newSkills])
                    injuryNiggling = row[12]
                    injuryMA = row[15]
                    injuryST = row[16]
                    injuryAG = row[17]
                    injuryAV = row[18]
                    if injuryNiggling:
                        skills.append("Niggling")
                    if injuryMA:
                        skills.append("-MA")
                    if injuryST:
                        skills.append("-ST")
                    if injuryAG:
                        skills.append("-AG")
                    if injuryAV:
                        skills.append("-AV")

                    if row[11]:
                        pass

                    status = "MNG" if row[11] else None
                    completions = row[20]
                    touchdowns = row[21]
                    interceptions = row[19]
                    casualties = row[22]
                    kills = row[23]
                    mvps = row[25]
                    players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])

                print "        %s" % teamName
                TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 6", players, teamValue, masterChef, bribes])

                if coachName not in COACHES:
                    coachFound = False
                    for coachNameKey, coachInfo in COACHES.iteritems():
                        if coachName in coachInfo["aka"]:
                            coachFound = True
                            break
                    if not coachFound:
                        print "ERROR: Coach '%s' not found!" % coachName
                        exit()
            print "Loaded %s rosters in %s" % (len(sheetNames), t.elapsed())
            totalRosters += len(sheetNames)



    MATCHES = []
    # Season 2 Matches
    if True:
        print
        print "Loading Season 2 Matches..."
        t.start()
        s3 = load_workbook("imports/S2 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s3.get_sheet_names()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s3[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            team1 = rows[2][1]
            team1TV = rows[2][25]
            team1InducementGP = rows[6][1] if rows[6][1] else 0
            team1Gate = rows[6][13]
            team1Fame = rows[6][25]
            team1Inducements = [rows[i][1] for i in range(9, 16)]
            team1Inducements.extend([rows[i][13] for i in range(9, 16)])
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Players = []
            for i in range(22, 53, 2):
                team1Players.append([rows[i][j] for j in range(3, 26, 2)])
                team1TDs += rows[i][5] if rows[i][5] else 0
                team1Cas += rows[i][9] if rows[i][9] else 0
                team1Kills += rows[i][11] if rows[i][11] else 0

            team1Winnings = rows[56][1]
            team1ExpMistakes = abs(rows[56][14]) if rows[56][14] else 0
            team1EndFanFactor = rows[56][27] if rows[56][27] else 0
            team1Purchases = rows[60][1]
            team1Notes = rows[60][19]

            team2 = rows[2][39]
            team2TV = rows[2][63]
            team2InducementGP = rows[6][39] if rows[6][39] else 0
            team2Gate = rows[6][51]
            team2Fame = rows[6][63]
            team2Inducements = [rows[i][39] for i in range(9, 16)]
            team2Inducements.extend([rows[i][52] for i in range(9, 16)])
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Players = []
            for i in range(22, 53, 2):
                team2Players.append([rows[i][j] for j in range(41, 64, 2)])
                team2TDs += rows[i][43] if rows[i][43] else 0
                team2Cas += rows[i][47] if rows[i][47] else 0
                team2Kills += rows[i][49] if rows[i][49] else 0

            team2Winnings = rows[56][39]
            team2ExpMistakes = abs(rows[56][52]) if rows[56][52] else 0
            team2EndFanFactor = rows[56][65] if rows[56][65] else 0
            team2Purchases = rows[60][39]
            team2Notes = rows[60][57]

            try:
                match = {
                    "type": "Regular",
                    "date": None,
                    "season": "Season 2",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)

    # Season 3 Matches
    if True:
        print
        print "Loading Season 3 Matches..."
        t.start()
        s3 = load_workbook("imports/S3 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s3.get_sheet_names()
        sheetNames.reverse()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s3[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            matchDate = rows[1][17]
            team1 = rows[2][1]
            team1TV = rows[2][14]
            team1InducementGP = rows[5][1]
            team1Gate = rows[5][7]
            team1Fame = rows[7][10]
            team1Inducements = [rows[i][1] for i in range(6, 9)]
            if team1Inducements[-1] and "\n" in team1Inducements[-1]:
                inducements = team1Inducements.pop().split("\n")
                for inducement in inducements:
                    team1Inducements.append(inducement)
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Players = []
            for i in range(13, 29):
                team1Players.append([rows[i][j] for j in range(2, 14)])
                team1TDs += rows[i][3] if rows[i][3] else 0
                team1Cas += rows[i][5] if rows[i][5] else 0
                team1Kills += rows[i][6] if rows[i][6] else 0

            team1Winnings = rows[31][1]
            team1ExpMistakes = rows[31][7] if rows[31][7] else 0
            team1EndFanFactor = rows[31][14] if rows[31][14] else 0
            team1Purchases = rows[34][1]
            team1Notes = rows[34][11]

            team2 = rows[2][22]
            team2TV = rows[2][35]
            team2InducementGP = rows[5][22]
            team2Gate = rows[5][28]
            team2Fame = rows[7][31]
            team2Inducements = [rows[i][22] for i in range(6, 9)]
            if team2Inducements[-1] and "\n" in team2Inducements[-1]:
                inducements = team2Inducements.pop().split("\n")
                for inducement in inducements:
                    team2Inducements.append(inducement)
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Players = []
            for i in range(13, 29):
                team2Players.append([rows[i][j] for j in range(23, 35)])
                team2TDs += rows[i][24] if rows[i][24] else 0
                team2Cas += rows[i][26] if rows[i][26] else 0
                team2Kills += rows[i][27] if rows[i][27] else 0

            team2Winnings = rows[31][22]
            team2ExpMistakes = rows[31][28] if rows[31][28] else 0
            team2EndFanFactor = rows[31][35] if rows[31][35] else 0
            team2Purchases = rows[34][22]
            team2Notes = rows[34][32]

            try:
                match = {
                    "type": "Regular",
                    "date": matchDate,
                    "season": "Season 3",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)

    # Season 3.5 Matches
    if True:
        print
        print "Loading Season 3.5 Matches..."
        t.start()
        s3p5 = load_workbook("imports/S3.5 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s3p5.get_sheet_names()
        sheetNames.reverse()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s3p5[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            matchDate = rows[1][19]
            team1 = rows[1][1]
            team1TV = rows[1][13]
            team1InducementGP = rows[4][1]
            team1Gate = rows[8][12]
            team1Fame = rows[10][14]
            team1Inducements = [rows[i][1] for i in range(5, 11)]
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Fouls = 0
            team1Players = []
            for i in range(16, 32):
                team1Players.append([rows[i][j] for j in range(2, 15)])
                team1TDs += rows[i][3] if rows[i][3] else 0
                team1Cas += rows[i][5] if rows[i][5] else 0
                team1Kills += rows[i][6] if rows[i][6] else 0
                team1Fouls += rows[i][7] if rows[i][7] else 0

            team1Winnings = rows[33][1]
            team1ExpMistakes = rows[33][6] if str(rows[33][6]).isdigit() else 0
            team1EndFanFactor = rows[33][17] if rows[33][17] else 0
            team1Purchases = [rows[i][1] for i in range(35, 38)]
            team1Notes = rows[35][12]

            team2 = rows[1][25]
            team2TV = rows[1][37]
            team2InducementGP = rows[4][23]
            team2Gate = rows[8][34]
            team2Fame = rows[10][36]
            team2Inducements = [rows[i][23] for i in range(5, 11)]
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Fouls = 0
            team2Players = []
            for i in range(16, 32):
                team2Players.append([rows[i][j] for j in range(24, 37)])
                team2TDs += rows[i][24] if rows[i][24] else 0
                team2Cas += rows[i][26] if rows[i][26] else 0
                team2Kills += rows[i][27] if rows[i][27] else 0
                team2Fouls += rows[i][28] if rows[i][28] else 0

            team2Winnings = rows[33][23]
            team2ExpMistakes = rows[33][28] if str(rows[33][28]).isdigit() else 0
            team2EndFanFactor = rows[33][39] if rows[33][39] else 0
            team2Purchases = [rows[i][23] for i in range(35, 38)]
            team2Notes = rows[35][34]

            try:
                match = {
                    "type": "Regular",
                    "date": matchDate,
                    "season": "Season 3.5",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)

    # Season 4 Matches
    if True:
        print
        print "Loading Season 4 Matches..."
        t.start()
        s4 = load_workbook("imports/S4 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s4.get_sheet_names()
        sheetNames.reverse()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s4[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            matchDate = rows[1][19]
            team1 = rows[1][1]
            team1TV = rows[1][13]
            team1InducementGP = rows[4][1]
            team1Gate = rows[8][12]
            team1Fame = rows[10][14]
            team1Inducements = [rows[i][1] for i in range(5, 11)]
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Fouls = 0
            team1Players = []
            for i in range(16, 32):
                team1Players.append([rows[i][j] for j in range(2, 15)])
                team1TDs += rows[i][3] if rows[i][3] else 0
                team1Cas += rows[i][5] if rows[i][5] else 0
                team1Kills += rows[i][6] if rows[i][6] else 0
                team1Fouls += rows[i][7] if rows[i][7] else 0

            team1Winnings = rows[33][1]
            team1ExpMistakes = rows[33][6] if str(rows[33][6]).isdigit() else 0
            team1EndFanFactor = rows[33][17] if rows[33][17] else 0
            team1Purchases = [rows[i][1] for i in range(35, 38)]
            team1Notes = rows[35][12]

            team2 = rows[1][25]
            team2TV = rows[1][37]
            team2InducementGP = rows[4][23]
            team2Gate = rows[8][34]
            team2Fame = rows[10][36]
            team2Inducements = [rows[i][23] for i in range(5, 11)]
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Fouls = 0
            team2Players = []
            for i in range(16, 32):
                team2Players.append([rows[i][j] for j in range(24, 37)])
                team2TDs += rows[i][24] if rows[i][24] else 0
                team2Cas += rows[i][26] if rows[i][26] else 0
                team2Kills += rows[i][27] if rows[i][27] else 0
                team2Fouls += rows[i][28] if rows[i][28] else 0

            team2Winnings = rows[33][23]
            team2ExpMistakes = rows[33][28] if str(rows[33][28]).isdigit() else 0
            team2EndFanFactor = rows[33][39] if rows[33][39] else 0
            team2Purchases = [rows[i][23] for i in range(35, 38)]
            team2Notes = rows[35][34]

            try:
                match = {
                    "type": "Regular",
                    "date": matchDate,
                    "season": "Season 4",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)

    # Season 5 Matches
    if True:
        print
        print "Loading Season 5 Matches..."
        t.start()
        s5 = load_workbook("imports/S5 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s5.get_sheet_names()[1:]
        sheetNames.reverse()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s5[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            matchDate = rows[1][19]
            team1 = rows[1][1]
            team1TV = rows[1][13]
            team1InducementGP = rows[4][1]
            team1Gate = rows[8][12]
            team1Fame = rows[10][14]
            team1Inducements = [rows[i][1] for i in range(5, 11)]
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Fouls = 0
            team1Players = []
            for i in range(16, 32):
                team1Players.append([rows[i][j] for j in range(2, 15)])
                team1TDs += rows[i][3] if rows[i][3] else 0
                team1Cas += rows[i][5] if rows[i][5] else 0
                team1Kills += rows[i][6] if rows[i][6] else 0
                team1Fouls += rows[i][7] if rows[i][7] else 0

            team1Winnings = rows[33][1]
            team1ExpMistakes = rows[33][6] if str(rows[33][6]).isdigit() else 0
            team1EndFanFactor = rows[33][17] if rows[33][17] else 0
            team1Purchases = [rows[i][1] for i in range(35, 38)]
            team1Notes = rows[35][12]

            team2 = rows[1][25]
            team2TV = rows[1][37]
            team2InducementGP = rows[4][23]
            team2Gate = rows[8][34]
            team2Fame = rows[10][36]
            team2Inducements = [rows[i][23] for i in range(5, 11)]
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Fouls = 0
            team2Players = []
            for i in range(16, 32):
                team2Players.append([rows[i][j] for j in range(24, 37)])
                team2TDs += rows[i][24] if rows[i][24] else 0
                team2Cas += rows[i][26] if rows[i][26] else 0
                team2Kills += rows[i][27] if rows[i][27] else 0
                team2Fouls += rows[i][28] if rows[i][28] else 0

            team2Winnings = rows[33][23]
            team2ExpMistakes = rows[33][28] if str(rows[33][28]).isdigit() else 0
            team2EndFanFactor = rows[33][39] if rows[33][39] else 0
            team2Purchases = [rows[i][23] for i in range(35, 38)]
            team2Notes = rows[35][34]

            try:
                match = {
                    "type": "Regular",
                    "date": matchDate,
                    "season": "Season 5",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)

    # Season 6 Matches
    if True:
        print
        print "Loading Season 6 Matches..."
        t.start()
        s6 = load_workbook("imports/S6 Matches.xlsx", data_only=True, read_only=True)
        print "    Sheet loaded in %s! Now loading matches..." % t.elapsed()
        t.start()
        sheetNames = s6.get_sheet_names()[1:]
        sheetNames.reverse()
        for sheetName in sheetNames:
            print "        %s" % sheetName
            matchSheet = s6[sheetName]
            rows = []
            for row in matchSheet.rows:
                rows.append([cell.value for cell in row])

            matchDate = rows[1][19]
            team1 = rows[1][1]
            team1TV = rows[1][13]
            team1InducementGP = rows[4][1]
            team1Gate = rows[8][12]
            team1Fame = rows[10][14]
            team1Inducements = [rows[i][1] for i in range(5, 11)]
            team1TDs = 0
            team1Cas = 0
            team1Kills = 0
            team1Fouls = 0
            team1Players = []
            for i in range(16, 32):
                team1Players.append([rows[i][j] for j in range(2, 15)])
                team1TDs += rows[i][3] if rows[i][3] else 0
                team1Cas += rows[i][5] if rows[i][5] else 0
                team1Kills += rows[i][6] if rows[i][6] else 0
                team1Fouls += rows[i][7] if rows[i][7] else 0

            team1Winnings = rows[33][1]
            team1ExpMistakes = rows[33][6] if str(rows[33][6]).isdigit() else 0
            team1EndFanFactor = rows[33][17] if rows[33][17] else 0
            team1Purchases = [rows[i][1] for i in range(35, 38)]
            team1Notes = rows[35][12]

            team2 = rows[1][25]
            team2TV = rows[1][37]
            team2InducementGP = rows[4][23]
            team2Gate = rows[8][34]
            team2Fame = rows[10][36]
            team2Inducements = [rows[i][23] for i in range(5, 11)]
            team2TDs = 0
            team2Cas = 0
            team2Kills = 0
            team2Fouls = 0
            team2Players = []
            for i in range(16, 32):
                team2Players.append([rows[i][j] for j in range(24, 37)])
                team2TDs += rows[i][24] if rows[i][24] else 0
                team2Cas += rows[i][26] if rows[i][26] else 0
                team2Kills += rows[i][27] if rows[i][27] else 0
                team2Fouls += rows[i][28] if rows[i][28] else 0

            team2Winnings = rows[33][23]
            team2ExpMistakes = rows[33][28] if str(rows[33][28]).isdigit() else 0
            team2EndFanFactor = rows[33][39] if rows[33][39] else 0
            team2Purchases = [rows[i][23] for i in range(35, 38)]
            team2Notes = rows[35][34]

            try:
                match = {
                    "type": "Regular",
                    "date": matchDate,
                    "season": "Season 6",
                    "teams": [
                        {
                            "name": team1.strip(),
                            "tv": team1TV,
                            "inducementsGP": int(team1InducementGP.replace("gp", "") if type(team1InducementGP) is str else team1InducementGP),
                            "gate": int(team1Gate),
                            "fame": int(team1Fame) if team1Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team1Inducements)),
                            "tds": int(team1TDs),
                            "cas": int(team1Cas),
                            "kills": int(team1Kills),
                            "players": team1Players,
                            "winnings": int(team1Winnings),
                            "expMistakes": int(team1ExpMistakes),
                            "fanFactor": int(team1EndFanFactor),
                            "purchases": team1Purchases,
                            "notes": team1Notes
                        },
                        {
                            "name": team2.strip(),
                            "tv": int(team2TV),
                            "inducementsGP": int(team2InducementGP.replace("gp", "") if type(team2InducementGP) is str else team2InducementGP),
                            "gate": int(team2Gate),
                            "fame": int(team2Fame) if team2Fame else 0,
                            "inducements": list(filter(lambda x: x is not None, team2Inducements)),
                            "tds": int(team2TDs),
                            "cas": int(team2Cas),
                            "kills": int(team2Kills),
                            "players": team2Players,
                            "winnings": int(team2Winnings),
                            "expMistakes": int(team2ExpMistakes),
                            "fanFactor": int(team2EndFanFactor),
                            "purchases": team2Purchases,
                            "notes": team2Notes
                        }
                    ]
                }
                MATCHES.append(match)
            except Exception as ex:
                pass
        print "Loaded %s matches in %s" % (len(sheetNames), t.elapsed())
        totalMatches += len(sheetNames)


    print "Loaded spreadsheets in %s. Importing into database..." % t1.elapsed()
    t.start()

    cnx = mysql.connector.connect(user=settings.mysql_username, password=settings.mysql_password, host=settings.mysql_server, database=settings.mysql_db)
    cnx.autocommit = True
    cursor = cnx.cursor()

    print
    print "Skill Types"
    skillTypeIDs = {}
    for skillType in SKILL_TYPES:
        print "    %s" % skillType
        cursor.execute("INSERT INTO skill_type VALUES(NULL,%s)", (skillType,))
        skillTypeID = cursor.lastrowid
        skillTypeIDs[skillType] = skillTypeID

    skillIDs = {}
    skillAliases = {
        "Blood Lust": ["Blood Blust"],
        "Throw Team-Mate": ["Throw Teammate", "TTM", "Throw Team-mate"],
        "Bone-head": ["Bonehead", "Bone Head", "Bone-headed", "Bone-Head"],
        "Side Step": ["Sidestep"],
        "Claws": ["Claw"],
        "Dodge": ["dodge"],
        "Leap": ["leap"],
        "Block": ["block"],
        "Secret Weapon": ["Secret Weapons"],
        "Bombardier": ["Bombadier"],
        "Thick Skull": ["Thick Skulld"],
        "Dump-off": ["Dump-Off"],
        "Prehensile Tail": ["Prehensail Tail", "Prehensal Tail"],
        "Mighty Blow": ["Might Blow"],
        "Nurgle's Rot": ["Nurgles Rot"],
        "Foul Appearance": ["Foul Apperance"],
        "+MA": ["+MV", "Movement"],
        "+AG": ["Agility"],
        "+AV": ["AV"]
    }
    print
    print "Skills"
    for skill in SKILLS:
        name = skill[0]
        print "    %s" % name
        skillType = skill[1]
        skillTypeID = skillTypeIDs[skillType]
        desc = skill[2]
        cursor.execute("INSERT INTO skill VALUES(NULL,%s,%s,%s)", (name, skillTypeID, desc))
        skillID = cursor.lastrowid
        skillIDs[name] = skillID

        aliases = getAliases(name, skillAliases)
        for alias in aliases:
            skillIDs[alias] = skillID

    raceAliases = {
        "Chaos Chosen": ["Chaos"],
        "Elven Union": ["Elf"],
        "Khemri Tomb Kings": ["Khemri"],
        "Necromantic Horror": ["Necromantic"],
        "Shambling Undead": ["Undead"],
        "Underworld Denizens": ["Underworld"]
    }
    raceIDs = {}
    print
    print "Races"
    for race in RACES:
        name = race[0]
        print "    %s" % name
        desc = race[1]
        cursor.execute("INSERT INTO race VALUES(NULL,%s,%s)", (name, desc))
        raceID = cursor.lastrowid
        raceIDs[name] = raceID

        aliases = getAliases(name, raceAliases)
        for alias in aliases:
            raceIDs[alias] = raceID

    coachIDs = {}
    print
    print "Coaches"
    for coachName, coachInfo in COACHES.iteritems():
        print "    %s" % coachName
        coachEmail = coachInfo["email"]
        cursor.execute("INSERT INTO coach VALUES(NULL,%s,%s,NULL,NULL)", (coachName, coachEmail))
        coachID = cursor.lastrowid
        coachIDs[coachName] = coachID

        for coachAlias in coachInfo["aka"]:
            coachIDs[coachAlias] = coachID

    trophyIDs = {}
    print
    print "Trophies"
    for trophy in TROPHIES:
        name = trophy[0]
        img = trophy[1]
        print "    %s" % name
        cursor.execute("INSERT INTO trophy VALUES(NULL,%s,%s)", (name, img))
        trophyID = cursor.lastrowid
        trophyIDs[name] = trophyID

    seasonIDs = {}
    print
    print "Seasons"
    for season in SEASONS:
        name = season[0]
        print "    %s" % name
        startDate = season[1]
        endDate = season[2]
        trophyName = season[3]
        trophyID = trophyIDs[trophyName]
        cursor.execute("INSERT INTO season VALUES(NULL,%s,%s,%s,NULL,%s)", (name, startDate, endDate, trophyID))
        seasonID = cursor.lastrowid
        seasonIDs[name] = seasonID

    playerTypeAliases = {
        "Amazon": {
            "Tribal Linewoman": ["Linewoman"],
            "Eagle Warrior Thrower": ["Thrower"],
            "Piranha Warrior Catcher": ["Catcher"],
            "Koka Kalim Blitzer": ["Blitzer"]
        },
        "Chaos Dwarf": {
            "Chaos Dwarf": ["Blocker"],
            "Enslaved Minotaur": ["Minotaur"]
        },
        "Chaos Pact": {
            "Human Renegade": ["Marauder"]
        },
        "Necromantic Horror": {
            "Ghoul Runner": ["Ghoul"],
            "Wight Blitzer": ["Wight"]
        },
        "Necromantic": {
            "Ghoul Runner": ["Ghoul"],
            "Wight Blitzer": ["Wight"]
        },
        "Norse": {
            "Berserker": ["Blitzer"],
            "Ulfwerenar": ["Werewolf"]
        },
        "Nurgle": {
            "Bloater": ["Warrior"],
            "Rotspawn": ["Beast"]
        },
        "Ogre": {
            "Runt": ["Snotling"]
        },
        "Orc": {
            "Black Orc": ["Black Orc Blocker"]
        },
        "Shambling Undead": {
            "Ghoul Runner": ["Ghoul"],
            "Wight Blitzer": ["Wight"]
        },
        "Undead": {
            "Ghoul Runner": ["Ghoul"],
            "Wight Blitzer": ["Wight"]
        }
    }
    playerTypeIDsByRaceID = {}
    playerTypeIDSkillIDs = {}
    print
    print "Players Types"
    # ["Underworld", 1, "Warpstone Troll", 110, 4, 5, 1, 9, [ "Loner", "Always Hungry", "Mighty Blow", "Really Stupid", "Regeneration", "Throw Team-Mate"], "SM", "GAP"]
    for playerType in PLAYER_TYPES:
        race = playerType[0]
        maxAmt = playerType[1]
        name = playerType[2]
        value = playerType[3]
        ma = playerType[4]
        st = playerType[5]
        ag = playerType[6]
        av = playerType[7]
        skills = playerType[8]
        normal = playerType[9]
        double = playerType[10]

        print "    %s: %s" % (race, name)
        cursor.execute("INSERT INTO player_type VALUES(NULL,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (name, maxAmt, ma, st, ag, av, 0, raceIDs[race], value, None))
        playerTypeID = cursor.lastrowid

        raceID = raceIDs[race]
        if raceID not in playerTypeIDsByRaceID:
            playerTypeIDsByRaceID[raceID] = {}
        playerTypeIDsByRaceID[raceID][name] = playerTypeID

        if race in playerTypeAliases:
            aliases = playerTypeAliases[race]
            for positional, posAliases in aliases.iteritems():
                if name == positional or name in posAliases:
                    for posAlias in posAliases:
                        playerTypeIDsByRaceID[raceID][posAlias] = playerTypeID

        playerTypeIDSkillIDs[playerTypeID] = []
        for skill in skills:
            skillID = skillIDs[skill]
            cursor.execute("INSERT INTO player_type_skill VALUES(%s,%s)", (playerTypeID, skillID))
            playerTypeIDSkillIDs[playerTypeID].append(skillID)

        for initial in normal:
            for skillType in skillTypeIDs.keys():
                if skillType.startswith(initial):
                    cursor.execute("INSERT INTO player_type_skill_type_normal VALUES(%s,%s)", (playerTypeID, skillTypeIDs[skillType]))

        for initial in double:
            for skillType in skillTypeIDs.keys():
                if skillType.startswith(initial):
                    cursor.execute("INSERT INTO player_type_skill_type_double VALUES(%s,%s)", (playerTypeID, skillTypeIDs[skillType]))


    # ["Willow Rosebark", ["Amazon", "Halfling", "Wood Elf"], ["Loner", "Dauntless", "Side Step", "Thick Skull"], 150, 5, 4, 3, 8, ""]
    print
    print "Star Players"
    starPlayerIDsByRaceID = {}
    for starPlayer in STAR_PLAYERS:
        name = starPlayer[0]
        races = starPlayer[1]
        skills = starPlayer[2]
        value = starPlayer[3]
        ma = starPlayer[4]
        st = starPlayer[5]
        ag = starPlayer[6]
        av = starPlayer[7]
        desc = starPlayer[8]

        if desc == "":
            desc = None

        if "All Except" in races:
            newRaces = raceIDs.keys()
            for race in races[1:]:
                raceID = raceIDs[race]
                newRaces.remove(race)
                for race2 in races[1:]:
                    if raceIDs[race2] == raceID and race2 in newRaces:
                        newRaces.remove(race2)
            races = newRaces


        addedRaceIDs = []
        for race in races:
            raceID = raceIDs[race]
            if raceID in addedRaceIDs:
                continue
            addedRaceIDs.append(raceID)

            print "    %s: %s" % (name, race)
            cursor.execute("INSERT INTO player_type VALUES(NULL,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", (name, 1, ma, st, ag, av, 1, raceID, value, desc))
            starPlayerID = cursor.lastrowid
            if raceID not in starPlayerIDsByRaceID:
                starPlayerIDsByRaceID[raceID] = {}
            starPlayerIDsByRaceID[raceID][name] = starPlayerID

            for skill in skills:
                cursor.execute("INSERT INTO player_type_skill VALUES(%s,%s)", (starPlayerID, skillIDs[skill]))

    teamNameAliases = {
        "Riverdale Ravagers": ["Alliance of the Old Gods", "Rivendale Ravagers"],
        "Run! It's The PO PO": ["Watch Out For The PO PO", "Watch Out For The POPO", "Run it's the Po Po", "Run its the PO PO", "Watch out for the POPO", "Watch out for the PO PO"],
        "Gold Diggers": ["Golddiggers", "Karak Barak Varr Gold Diggers", "GoldDiggers"],
        "Blood Mountain Berserkers": ["Blood Mountain Berzerkers"],
        "The DodgeBall All-Stars": ["Dodgeball All-Stars", "The Dodgeball All-Stars"],
        "The Flying Circus": ["Flying Circus"]
    }
    teamIDsBySeasonID = {}
    playerIDs = {}
    playerIDsBySeasonIDAndTeamAndNumber = {}
    playerIDPlayerTypeIDs = {}
    raceIDsByTeamID = {}
    season5ID = None
    print
    print "Teams"
    for team in TEAMS:
        # TEAMS.append([teamName, coachName, raceName, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, "Season 3", players, teamValue])
        teamName = team[0].strip()
        seasonName = team[9]
        print "    %s: %s" % (seasonName, teamName)

        coachName = team[1].strip()
        raceName = team[2].strip()
        treasury = int(team[3])
        value = int(team[11])
        rerolls = int(team[4]) if team[4] else 0
        fanFactor = int(team[5])
        assistantCoaches = int(team[6])
        cheerleaders = int(team[7])
        apothecary = int(team[8])
        players = team[10]
        masterChef = int(team[12])
        bribes = int(team[13])

        coachID = coachIDs[coachName]
        raceID = raceIDs[raceName]
        seasonID = seasonIDs[seasonName]

        if seasonName == "Season 5":
            season5ID = seasonID

        if seasonID not in teamIDsBySeasonID:
            teamIDsBySeasonID[seasonID] = {}

        prevTeamID = None
        if len(seasonIDs) > 1:
            if seasonID != season5ID: # Season 5 restarted all teams
                prevSeasonID = seasonID - 1
                
                minSeasonID = 1
                if seasonID > season5ID and season5ID is not None:
                    minSeasonID = season5ID

                while prevSeasonID >= minSeasonID and prevSeasonID > 0:
                    prevSeasonTeams = teamIDsBySeasonID[prevSeasonID]
                    
                    if teamName in prevSeasonTeams:
                        prevTeamID = prevSeasonTeams[teamName]
                        break

                    teamAliases = getAliases(teamName, teamNameAliases)
                    for alias in teamAliases:
                        if alias in prevSeasonTeams:
                            prevTeamID = prevSeasonTeams[alias]
                            break

                    if prevTeamID:
                        break

                    prevSeasonID -= 1

        if treasury > 9999:
            treasury /= 1000
        if value > 9999:
            value /= 1000

        row = (teamName, coachID, raceID, seasonID, prevTeamID, value, treasury, rerolls, fanFactor, assistantCoaches, cheerleaders, apothecary, masterChef, bribes)
        cursor.execute("INSERT IGNORE INTO team VALUES(NULL," + ",".join(["%s" for x in row]) + ")", row)
        teamID = cursor.lastrowid
        teamIDsBySeasonID[seasonID][teamName] = teamID
        raceIDsByTeamID[teamID] = raceID

        teamAliases = getAliases(teamName, teamNameAliases)
        for alias in teamAliases:
            teamIDsBySeasonID[seasonID][alias] = teamID

        for i, player in enumerate(players):
            # players.append([playerName, playerPosition, ma, st, ag, av, skills, status, completions, touchdowns, interceptions, casualties, kills, mvps])
            number = i + 1
            isStarPlayer = 0
            playerName = player[0].strip() if player[0] else None
            playerPosition = player[1].strip() if player[1] else None
            
            playerTypeID = None
            if playerPosition:
                if playerPosition.startswith("Star: "):
                    isStarPlayer = 1
                    starPlayerName = playerPosition.replace("Star: ", "")
                    starPlayerID = starPlayerIDsByRaceID[raceID][starPlayerName]
                    playerTypeID = starPlayerID
                    playerName = starPlayerName
                else:
                    playerTypeID = playerTypeIDsByRaceID[raceID][playerPosition]

            if not playerName and playerPosition:
                playerName = "Player %s" % number

            print "        (%s/%s): %s" % (number, len(players), playerName)
            if not playerPosition:
                continue

            if playerName.lower() == "journeyman":
                continue
            
            ma = player[2]
            st = player[3]
            ag = player[4]
            av = player[5]

            skills = player[6]
            mng = "1" if player[7] == "MNG" else "0"
            niggling = "1" if "Niggling" in skills else "0"
            injCount = collections.Counter(skills)
            maInjuries = injCount["-MA"] if "-MA" in injCount else 0
            stInjuries = injCount["-ST"] if "-ST" in injCount else 0
            agInjuries = injCount["-AG"] if "-AG" in injCount else 0
            avInjuries = injCount["-AV"] if "-AV" in injCount else 0
            skillsNoInjuries = set(skills) - set(["-MA", "-ST", "-AG", "-AV", "Niggling"])

            completions = player[8] if player[8] else 0
            touchdowns = player[9] if player[9] else 0
            interceptions = player[10] if player[10] else 0
            casualties = player[11] if player[11] else 0
            kills = player[12] if player[12] else 0
            mvps = player[13] if player[13] else 0

            prevPlayerID = None
            if isStarPlayer == 0:
                if prevTeamID is not None and prevSeasonID > 0:
                    playerReportsByNumber = playerIDsBySeasonIDAndTeamAndNumber[prevSeasonID][prevTeamID]
                    for prevPlayerReport in playerReportsByNumber.values():
                        prevPlayerName = prevPlayerReport[2]
                        if playerName.strip().lower() == prevPlayerName.strip().lower():
                            prevPlayerID = prevPlayerReport[0]
                            break

            playerReport = (number, playerName, teamID, playerTypeID, prevPlayerID, mng, niggling, maInjuries, stInjuries, agInjuries, avInjuries, interceptions, completions, touchdowns, casualties, kills, mvps, isStarPlayer)
            cursor.execute("INSERT INTO player VALUES(NULL," + ",".join(["%s" for pr in playerReport]) + ")", playerReport)
            playerID = cursor.lastrowid
            playerIDPlayerTypeIDs[playerID] = playerTypeID
            playerReport = (playerID,) + playerReport

            if playerID <= 0:
                print "ERROR: Player not imported!"

            if isStarPlayer == 0:
                for skillName in skillsNoInjuries:
                    if skillName not in skillIDs:
                        pass
                    else:
                        skillID = skillIDs[skillName]
                        playerTypeSkills = playerTypeIDSkillIDs[playerTypeID]
                        if skillID not in playerTypeSkills:
                            cursor.execute("INSERT INTO player_skill VALUES(%s,%s)", (playerID, skillID))

            if teamName not in playerIDs:
                playerIDs[teamName] = []
            playerIDs[teamName].append(playerReport)

            teamAliases = getAliases(teamName, teamNameAliases)
            for alias in teamAliases:
                if alias not in playerIDs:
                    playerIDs[alias] = []
                playerIDs[alias].append(playerReport)

            if seasonID not in playerIDsBySeasonIDAndTeamAndNumber:
                playerIDsBySeasonIDAndTeamAndNumber[seasonID] = {}
            if teamID not in playerIDsBySeasonIDAndTeamAndNumber[seasonID]:
                playerIDsBySeasonIDAndTeamAndNumber[seasonID][teamID] = {}
            playerIDsBySeasonIDAndTeamAndNumber[seasonID][teamID][number] = playerReport

    print
    print "Match Types"
    matchTypeIDs = {}
    for matchType in MATCH_TYPES:
        print "    %s" % matchType
        cursor.execute("INSERT INTO match_type VALUES(NULL,%s)", (matchType,))
        matchTypeID = cursor.lastrowid
        matchTypeIDs[matchType] = matchTypeID

    print
    print "Inducements"
    inducementAliases = {
        "Marketing Blitz": ["Fan Factor"],
        "Extra Team Training": ["Team Re-roll"],
        "Bloodweiser Babe": ["Budweiser Babe", "Babe", "Bloodwiser Babe", "Bloodweiser Babes"]
    }
    inducementIDs = {}
    inducementIDsByRaceID = {}
    # ["Bloodweiser Babe", 2, {"All": 50}, "The team
    for inducement in INDUCEMENTS:
        name = inducement[0]
        maxAmt = inducement[1]
        priceForRaces = inducement[2]
        desc = inducement[3]

        for race, price in priceForRaces.iteritems():
            raceID = None
            if race != "All" and race != "Otherwise":
                raceID = raceIDs[race]

            print "    %s: %s" % (name, race)
            cursor.execute("INSERT INTO inducement VALUES(NULL,%s,%s,%s,%s,%s)", (name, maxAmt, price, raceID, desc))
            inducementID = cursor.lastrowid

            if raceID:
                if raceID not in inducementIDsByRaceID:
                    inducementIDsByRaceID[raceID] = {}
                inducementIDsByRaceID[raceID][name] = inducementID

                if name in inducementAliases:
                    for alias in inducementAliases[name]:
                        inducementIDsByRaceID[raceID][alias] = inducementID
            else:
                inducementIDs[name] = inducementID
                if name in inducementAliases:
                    for alias in inducementAliases[name]:
                        inducementIDs[alias] = inducementID
                        # ["Cheerleader", 10, ["All"], "Most Blood Bow
    print
    print "Purchases"
    purchaseIDs = {}
    purchaseIDsByRaceID = {}
    # ["Apothecary", {"All":50, "Exceptions":["Khemri", "Necromantic", "Nurgle", "Undead"]}, "An Apothecary is a
    for purchase in PURCHASES:
        name = purchase[0]
        maxAmt = purchase[1]
        raceCosts = purchase[2]
        desc = purchase[3]

        for raceNameCost, cost in raceCosts.iteritems():
            if raceNameCost == "All":
                if "Exceptions" in raceCosts:
                    raceNames = raceIDs.keys()
                    exceptRaceNames = raceCosts["Exceptions"]
                    for exceptRaceName in exceptRaceNames:
                        exceptRaceID = raceIDs[exceptRaceName]
                        raceNames.remove(exceptRaceName)
                        for race, raceID  in raceIDs.iteritems():
                            if raceID == exceptRaceID and race in raceNames:
                                raceNames.remove(race)

                    addedRaceIDs = []
                    for raceName in raceNames:
                        raceID = raceIDs[raceName]
                        if raceID in addedRaceIDs:
                            continue
                        addedRaceIDs.append(raceID)

                        print "    %s: %s (%s,000 gp)" % (name, raceName, cost)
                        cursor.execute("INSERT INTO purchase VALUES(NULL,%s,%s,%s,%s,%s)", (name, maxAmt, cost, raceID, desc))
                        purchaseID = cursor.lastrowid
                        if raceID not in purchaseIDsByRaceID:
                            purchaseIDsByRaceID[raceID] = {}
                        purchaseIDsByRaceID[raceID][name] = purchaseID
                    break
                else:
                    print "    All: %s (%s,000 gp)" % (name, cost)
                    cursor.execute("INSERT INTO purchase VALUES(NULL,%s,%s,%s,NULL,%s)", (name, maxAmt, cost, desc))
                    purchaseID = cursor.lastrowid
                    purchaseIDs[name] = purchaseID
            elif raceNameCost != "Exceptions":
                raceName = raceNameCost
                print "    %s: %s (%s,000 gp)" % (name, raceName, cost)
                raceID = raceIDs[raceName]
                cursor.execute("INSERT INTO purchase VALUES(NULL,%s,%s,%s,%s,%s)", (name, maxAmt, cost, raceID, desc))
                purchaseID = cursor.lastrowid
                if raceID not in purchaseIDsByRaceID:
                    purchaseIDsByRaceID[raceID] = {}
                purchaseIDsByRaceID[raceID][name] = purchaseID

                if "Otherwise" in raceCosts:
                    cost = raceCosts["Otherwise"]
                    raceID = None

                    print "    %s: All Other Races (%s,000 gp)" % (name, cost)
                    cursor.execute("INSERT INTO purchase VALUES(NULL,%s,%s,%s,%s,%s)", (name, maxAmt, cost, raceID, desc))
                    purchaseID = cursor.lastrowid
                    purchaseIDs[name] = purchaseID
                    break

    print
    print "Decks"
    deckIDs = {}
    deckNamesByDeckID = {}
    for deck in DECKS:
        name = deck[0]
        cost = deck[1]
        print "    %s" % name
        cursor.execute("INSERT INTO deck VALUES(NULL,%s,%s)", (name, cost))
        deckID = cursor.lastrowid
        deckIDs[name] = deckID
        deckNamesByDeckID[deckID] = name

    print
    print "Cards"
    cardIDs = {}
    deckIDsByCardID = {}
    for card in CARDS:
        name = card[0]
        aka = card[1]
        deckName = card[2]
        desc = card[3]
        timing = card[4]
        effect = card[5]
        print "    %s" % name

        deckID = deckIDs[deckName]
        cursor.execute("INSERT IGNORE INTO card VALUES(NULL,%s,%s,%s,%s,%s,%s)", (deckID, name, aka, desc, timing, effect))
        cardID = cursor.lastrowid
        cardIDs[name] = cardID
        deckIDsByCardID[cardID] = deckID

    print
    print "Matches"
    matchNum = 0
    curSeason = None
    for match in MATCHES:
        if curSeason != match["season"]:
            curSeason = match["season"]
            matchNum = 1

        team1 = match["teams"][0]
        team2 = match["teams"][1]
        print "    %s: Match %s - \"%s\" vs. \"%s\"" % (curSeason, matchNum, team1["name"], team2["name"])

        seasonID = seasonIDs[curSeason]
        matchTypeID = matchTypeIDs[match["type"]]

        team1Name = team1["name"]
        team1ID = teamIDsBySeasonID[seasonID][team1Name] if team1Name in teamIDsBySeasonID[seasonID] else None
        team1Aliases = getAliases(team1Name, teamNameAliases)
        if not team1ID:
            for alias in team1Aliases:
                if alias in teamIDsBySeasonID[seasonID]:
                    team1ID = teamIDsBySeasonID[seasonID][team1Name]
                    break

        if not team1ID:
            print "Error: Couldn't find team1ID for team '%s'" % team1Name

        team2Name = team2["name"]
        team2ID = teamIDsBySeasonID[seasonID][team2Name] if team2Name in teamIDsBySeasonID[seasonID] else None
        if not team2ID:
            team2Aliases = getAliases(team2["name"], teamNameAliases)
            for alias in team2Aliases:
                if alias in teamIDsBySeasonID[seasonID]:
                    team2ID = teamIDsBySeasonID[seasonID][team2["name"]]
                    break

        if not team2ID:
            print "Error: Couldn't find team2ID for team '%s'" % team2["name"]

        description = ""
        if team1["notes"]:
            description = "%s: %s" % (team1["name"], team1["notes"])
        if team2["notes"]:
            if team1["notes"]:
                description += "\n" + "%s: %s" % (team2["name"], team2["notes"])
            else:
                description = "%s: %s" % (team2["name"], team2["notes"])

        if team1["tv"] > 1000:
            team1["tv"] /= 1000
        if team2["tv"] > 1000:
            team2["tv"] /= 1000

        if team1["inducementsGP"] > 1000:
            team1["inducementsGP"] /= 1000
        if team2["inducementsGP"] > 1000:
            team2["inducementsGP"] /= 1000

        if team1["gate"] > 1000:
            team1["gate"] /= 1000
            if team1["gate"] > 100:
                team1["gate"] /= 100
        if team2["gate"] > 1000:
            team2["gate"] /= 1000
            if team2["gate"] > 100:
                team2["gate"] /= 100

        if team1["winnings"] > 1000:
            team1["winnings"] /= 1000
        if team2["winnings"] > 1000:
            team2["winnings"] /= 1000

        if team1["expMistakes"] > 1000:
            team1["expMistakes"] /= 1000
        if team2["expMistakes"] > 1000:
            team2["expMistakes"] /= 1000

        row = (seasonID, match["date"], description,
               matchTypeID, matchTypeID,
               team1ID, team2ID,
               team1["tv"], team2["tv"],
               team1["inducementsGP"], team2["inducementsGP"],
               0, 0,
               team1["gate"], team2["gate"],
               team1["fame"], team2["fame"],
               team1["tds"], team2["tds"],
               team1["cas"], team2["cas"],
               team1["kills"], team2["kills"],
               team1["winnings"], team2["winnings"],
               team1["expMistakes"], team2["expMistakes"],
               team1["fanFactor"], team2["fanFactor"])
        insertQry = "INSERT IGNORE INTO `match` VALUES(NULL," + ",".join("%s" for entry in row) + ")"
        cursor.execute(insertQry, row)
        matchID = cursor.lastrowid

        seasonTeams = playerIDsBySeasonIDAndTeamAndNumber[seasonID]
        for team in match["teams"]:
            teamID = teamIDsBySeasonID[seasonID][team["name"]]
            raceID = raceIDsByTeamID[teamID]

            teamInducements = team["inducements"]
            for inducement in teamInducements:
                print "        Inducement: %s" % inducement
                inducementID = None
                playerTypeID = None
                deckID = None
                cardID = None
                amount = 1

                inducementSplit = inducement.split()
                if inducementSplit[0][0].isdigit() and inducementSplit[0].lower().endswith("x"):
                    amount = inducementSplit[0][0]
                    inducement = " ".join(inducementSplit[1:])

                if raceID in inducementIDsByRaceID:
                    raceInducementIDs = inducementIDsByRaceID[raceID]
                    if inducement in raceInducementIDs:
                        inducementID = raceInducementIDs[inducement]

                mercSkills = []
                if "Mercenary" in inducement:
                    inducement = inducement.replace("Mercenary", "").strip()
                    if " w/ " in inducement:
                        mercTemp = inducement.split("w/")
                        inducement = mercTemp[0].strip()
                        mercSkills = mercTemp[1].strip()
                        if len(mercSkills) > 0:
                            addDesc = "\n'%s' Mercenary %s also had the following skills: %s" % (team["name"], inducement, "/".join(mercSkills))
                            print "            %s" % addDesc.strip()
                            description += addDesc
                            cursor.execute("UPDATE `match` SET description=%s WHERE id=%s", (description, matchID))

                    if inducement in playerTypeIDsByRaceID[raceID]:
                        playerTypeID = playerTypeIDsByRaceID[raceID][inducement]
                elif "Journeyman" in inducement:
                    addDesc = "\n'%s' also had %sx %s" % (team["name"], amount, inducement)
                    print "            %s" % addDesc.strip()
                    description += addDesc
                    cursor.execute("UPDATE `match` SET description=%s WHERE id=%s", (description, matchID))
                    continue
                elif " Card" in inducement:
                    deckName = inducement.replace(" Card", "")
                    if deckName in deckIDs:
                        deckID = deckIDs[deckName]

                if inducementID is None:
                    if inducement in inducementIDs:
                        inducementID = inducementIDs[inducement]
                    elif inducement in starPlayerIDsByRaceID[raceID]:
                        playerTypeID = starPlayerIDsByRaceID[raceID][inducement]
                    elif inducement in cardIDs:
                        cardID = cardIDs[inducement]
                        deckID = deckIDsByCardID[cardID]
                        deckName = deckNamesByDeckID[deckID]
                        
                        addDesc = "\n'%s' drew the %s card named \"%s\"" % (team["name"], deckName, inducement)
                        print "            %s" % addDesc.strip()
                        description += addDesc
                        cursor.execute("UPDATE `match` SET description=%s WHERE id=%s", (description, matchID))
                    elif "petty cash" in inducement.lower():
                        pettyCash = inducement.lower().split()[0].replace("k", "")
                        teamNum = "1" if team["name"] == team1["name"] else "2"
                        cursor.execute("UPDATE `match` SET team" + teamNum + "_petty_cash=%s WHERE id=%s", (pettyCash, matchID))
                        continue

                if inducementID is None and playerTypeID is None and deckID is None:
                    print "            Couldn't find inducement '%s' Skipping..." % inducement
                    continue

                row = (matchID, teamID, inducementID, amount, playerTypeID, deckID, cardID)
                cursor.execute("INSERT INTO match_inducement VALUES(NULL," + ",".join("%s" for entry in row) + ")", row)

            teamPurchases = []
            if isinstance(team["purchases"], list):
                teamPurchases = team["purchases"]
            else:
                teamPurchases = team["purchases"].split("\n") if team["purchases"] else []
                
            for purchase in teamPurchases:
                if purchase == "" or purchase is None:
                    continue

                print "        Purchase: %s" % purchase
                purchaseID = None
                playerTypeID = None
                amount = 1

                purchaseSplit = purchase.split()
                if purchaseSplit[0][0].isdigit() and purchaseSplit[0].lower().endswith("x"):
                    amount = purchaseSplit[0][0]
                    purchase = " ".join(purchaseSplit[1:])

                if raceID in purchaseIDsByRaceID:
                    racePurchaseIDs = purchaseIDsByRaceID[raceID]
                    if purchase in racePurchaseIDs:
                        purchaseID = racePurchaseIDs[purchase]

                if purchaseID is None:
                    if purchase in playerTypeIDsByRaceID[raceID]:
                        playerTypeID = playerTypeIDsByRaceID[raceID][purchase]
                    elif purchase in purchaseIDs:
                        purchaseID = purchaseIDs[purchase]

                if purchaseID is None and playerTypeID is None:
                    print "            Couldn't find purchase '%s' Skipping..." % purchase
                    continue

                row = (matchID, teamID, purchaseID, amount, playerTypeID)
                cursor.execute("INSERT INTO match_purchase VALUES(NULL," + ",".join("%s" for entry in row) + ")", row)

            seasonTeamPlayers = seasonTeams[teamID]
            teamPlayers = playerIDs[team["name"]]
            for i, player in enumerate(team["players"]):
                playerNum = i + 1
                playerFromTeam = None
                if playerNum in seasonTeamPlayers:
                    playerFromTeam = seasonTeamPlayers[playerNum]
                else:
                    continue

                playerID = playerFromTeam[0]

                completions = player[0]
                touchdowns = player[1]
                interceptions = player[2]
                casualties = player[3]
                kills = player[4]
                fouls = None
                mvp = player[5]
                spp = player[6]
                mng = "1" if player[7] else None
                niggling = player[8]
                minusStat = player[9]
                dead = player[10]
                newSkill = player[11]

                if len(player) > 12:
                    fouls = player[5]
                    mvp = player[6]
                    spp = player[7]
                    mng = "1" if player[8] else None
                    niggling = player[9]
                    minusStat = player[10]
                    dead = player[11]
                    newSkill = player[12]

                mvInjury = None
                stInjury = None
                agInjury = None
                avInjury = None
                if minusStat is not None:
                    if minusStat == "ST":
                        stInjury = 1
                    elif minusStat == "MA" or minusStat == "MV":
                        mvInjury = 1
                    elif minusStat == "AG":
                        agInjury = 1
                    elif minusStat == "AV":
                        avInjury = 1
                    else:
                        print "ERROR: Cannot find injury for '%s'..." % minusStat

                if mng is not None:
                    mng = 1
                if niggling is not None:
                    niggling = 1
                    mng = 1

                row = (matchID, playerID, mng, niggling, mvInjury, stInjury, agInjury, avInjury, interceptions, completions, touchdowns, casualties, kills, fouls, mvp, dead)
                cursor.execute("INSERT INTO match_player VALUES(NULL," + ",".join("%s" for entry in row) + ")", row)

                if newSkill is not None:
                    newSkills = [s.strip() for s in newSkill.split(",")]
                    for skillName in newSkills:
                        if skillName != "Fired":
                            if skillName not in skillIDs:
                                pass
                            else:
                                newSkillID = skillIDs[skillName]
                                playerTypeID = playerIDPlayerTypeIDs[playerID]
                                playerTypeSkills = playerTypeIDSkillIDs[playerTypeID]
                                if newSkillID not in playerTypeSkills:
                                    cursor.execute("INSERT INTO match_player_skill VALUES(NULL,%s,%s,%s)", (matchID, playerID, newSkillID))
        matchNum += 1

    print
    print "Updating Season Winners..."
    for season in SEASONS:
        name = season[0]
        winnerTeamName = season[4]
        print "    %s: %s" % (name, winnerTeamName)
        if winnerTeamName:
            seasonID = seasonIDs[name]
            winnerTeamID = teamIDsBySeasonID[seasonID][winnerTeamName]
            cursor.execute("UPDATE season SET winner_team_id=%s WHERE id=%s", (winnerTeamID, seasonID))

    cursor.execute("INSERT INTO meta VALUES(NULL,%s,%s)", ("season.current", "9"))

    print
    print "Loaded data into database took %s" % t.elapsed()
    print "Importing %s rosters and %s matches took %s" % (totalRosters, totalMatches, t1.elapsed())

    print
    print "Done!"
